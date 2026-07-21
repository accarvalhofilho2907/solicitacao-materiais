from functools import wraps
from datetime import datetime

from flask import Blueprint, render_template, redirect, url_for, abort, flash, current_app, request, session
from flask_login import login_required, current_user, login_user

from .extensions import db
from .util import contem_busca
from .models import Solicitacao, LogSolicitacao
from .emails import enviar_email

almox_bp = Blueprint("almox", __name__, url_prefix="/almoxarifado")


def almox_required(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not (current_user.is_almox or current_user.is_admin):
            abort(403)
        return f(*args, **kwargs)
    return wrapper


@almox_bp.route("/chegadas")
@almox_required
def index():
    pendentes = (Solicitacao.query.filter_by(status="AGUARDANDO_CHEGADA")
                 .order_by(Solicitacao.prazo_recebimento).all())
    recentes = (Solicitacao.query.filter_by(status="CONCLUIDO")
                .order_by(Solicitacao.chegada_em.desc()).limit(10).all())
    return render_template("almox/index.html", pendentes=pendentes, recentes=recentes)


@almox_bp.route("/chegada/<int:sid>/editar-data", methods=["POST"])
@almox_required
def editar_data_chegada(sid):
    """Corrige a data de uma chegada já confirmada (item 129)."""
    s = db.session.get(Solicitacao, sid) or abort(404)
    nova_data = request.form.get("nova_data")
    if not nova_data or not s.chegada_em:
        flash("Informe uma data válida.", "danger")
        return redirect(url_for("almox.index"))
    try:
        d = datetime.strptime(nova_data, "%Y-%m-%d")
        s.chegada_em = d.replace(hour=s.chegada_em.hour, minute=s.chegada_em.minute, second=s.chegada_em.second)
        db.session.add(LogSolicitacao(solicitacao_id=s.id, autor_id=current_user.id,
                                      evento=f"Data de chegada corrigida para {d:%d/%m/%Y}"))
        db.session.commit()
        flash("Data de chegada atualizada.", "success")
    except ValueError:
        flash("Data inválida.", "danger")
    return redirect(url_for("almox.index"))


@almox_bp.route("/chegada/<int:sid>", methods=["POST"])
@almox_required
def marcar_chegada(sid):
    s = db.session.get(Solicitacao, sid) or abort(404)
    if s.status != "AGUARDANDO_CHEGADA":
        flash("Esta solicitação não está aguardando chegada.", "warning")
        return redirect(url_for("almox.index"))
    ja = s.quantidade_recebida or 0
    restante = max(s.quantidade - ja, 0)
    # quantidade informada nesta confirmação (padrão = restante)
    try:
        qtd = int(request.form.get("qtd_recebida") or restante)
    except ValueError:
        qtd = restante
    if qtd <= 0:
        flash("Informe uma quantidade recebida maior que zero.", "danger")
        return redirect(url_for("almox.index"))
    qtd = min(qtd, restante)
    total = ja + qtd
    s.quantidade_recebida = total

    # item 129 — permitir registrar com data diferente de hoje (ex.: chegada com atraso no registro)
    data_chegada_str = request.form.get("data_chegada")
    momento_chegada = datetime.utcnow()
    if data_chegada_str:
        try:
            data_escolhida = datetime.strptime(data_chegada_str, "%Y-%m-%d")
            momento_chegada = data_escolhida.replace(
                hour=momento_chegada.hour, minute=momento_chegada.minute, second=momento_chegada.second)
        except ValueError:
            pass

    if total >= s.quantidade:
        s.status = "CONCLUIDO"
        s.chegada_confirmada_por = current_user.id
        s.chegada_em = momento_chegada
        evento = f"Chegada confirmada — recebido {total} de {s.quantidade} (Concluído) — data: {momento_chegada:%d/%m/%Y}"
        assunto = f"Solicitação Nº {s.id} — material recebido (completo)"
        msg = f"Chegada completa: {total} de {s.quantidade} ({s.material})."
        flash(f"Chegada total confirmada. Solicitação Nº {s.id} concluída.", "success")
    else:
        evento = f"Chegada parcial — recebido {qtd} (acumulado {total} de {s.quantidade}) — data: {momento_chegada:%d/%m/%Y}"
        assunto = f"Solicitação Nº {s.id} — chegada parcial"
        msg = f"Chegada parcial: recebido {total} de {s.quantidade} ({s.material})."
        flash(f"Chegada parcial registrada: {total} de {s.quantidade}.", "success")
    db.session.add(LogSolicitacao(solicitacao_id=s.id, autor_id=current_user.id, evento=evento))
    db.session.commit()
    enviar_email([s.solicitante.email, current_app.config.get("ADMIN_EMAIL")], assunto,
                 f"A confirmação de chegada da solicitação Nº {s.id} foi registrada. {msg}")
    return redirect(url_for("almox.index"))


# ==================== MÓDULO ALMOXARIFADO (Chaves / Extintores / Colaboradores) ====================
from datetime import date
from .models import (Chave, Extintor, Colaborador, AlmoxLog, QuadroChave,
                     PapelColaborador, MovimentacaoChave, TAREFAS_COLABORADOR, TAREFAS_DICT,
                     TAREFAS_PERFIL, TAREFAS_GRUPOS,
                     InspecaoExtintor, PendenciaEtiqueta, CHECK_EXTINTOR, ITEM_ETIQUETA_EXTINTOR,
                     ProdutoAlmox, MovimentacaoMaterial, LocalAlmox, Fabricante,
                     NotaFiscalAlmox, NotificacaoAlmox, AjusteInventario,
                     EstoqueLocalizador, InstanciaItem)


def _qr_svg(texto, box=8, border=2):
    """Gera um QR Code como string SVG (sem dependência de imagem/Pillow).
    Usado para impressão de etiquetas de chaves/colaboradores/extintores."""
    try:
        import qrcode
        import qrcode.image.svg as _svg
        import io
        img = qrcode.make(texto, image_factory=_svg.SvgPathImage, box_size=box, border=border)
        buf = io.BytesIO()
        img.save(buf)
        return buf.getvalue().decode("utf-8")
    except Exception:
        # Fallback: caixa com o texto, caso a lib não esteja disponível
        return (f'<svg xmlns="http://www.w3.org/2000/svg" width="120" height="120">'
                f'<rect width="120" height="120" fill="#eee" stroke="#999"/>'
                f'<text x="60" y="64" font-size="10" text-anchor="middle">{texto}</text></svg>')

# Tópicos do módulo. liberado=True abre a tela; senão mostra "Em construção".
TOPICOS = [
    {"slug": "painel",      "nome": "Painel",                          "icone": "📊", "liberado": False},
    {"slug": "entrada",     "nome": "Entrada de material",             "icone": "⤵",  "liberado": False},
    {"slug": "devolucao",   "nome": "Devolução forçada",               "icone": "⏎",  "liberado": False},
    {"slug": "inventario",  "nome": "Inventário",                      "icone": "✓",  "liberado": False},
    {"slug": "ajuste",      "nome": "Ajuste de item",                  "icone": "✎",  "liberado": False},
    {"slug": "produto",     "nome": "Produto (Item/Classe/Atributo)",  "icone": "⬡",  "liberado": False},
    {"slug": "unidades",    "nome": "Unidades / validade / calibração","icone": "⌗",  "liberado": False},
    {"slug": "kit",         "nome": "Controle por Kit",                "icone": "◫",  "liberado": False},
    {"slug": "consulta",    "nome": "Consulta de estoque",             "icone": "🔎", "liberado": False},
    {"slug": "movimentacoes","nome": "Movimentações",                  "icone": "⇄",  "liberado": False},
    {"slug": "etiquetas",   "nome": "Etiquetas",                       "icone": "🏷", "liberado": False},
    {"slug": "qr",          "nome": "QR de colaboradores",             "icone": "▨",  "liberado": False},
    {"slug": "log",         "nome": "Log de ações",                    "icone": "📜", "liberado": False},
    {"slug": "coletor",     "nome": "Coletor",                         "icone": "📲", "liberado": True, "endpoint": "almox.coletor"},
    {"slug": "materiais",   "nome": "Material (estoque)",              "icone": "📦", "liberado": True, "endpoint": "almox.materiais"},
    {"slug": "mat_mov",     "nome": "Movimentações de material",       "icone": "📊", "liberado": True, "endpoint": "almox.materiais_mov"},
    {"slug": "chaves",      "nome": "Chaves",                          "icone": "🔑", "liberado": True, "endpoint": "almox.chaves"},
    {"slug": "relatorio_chaves","nome": "Relatório de chaves",         "icone": "📈", "liberado": True, "endpoint": "almox.relatorio_chaves"},
    {"slug": "extintores",  "nome": "Extintores",                      "icone": "🧯", "liberado": True, "endpoint": "almox.extintores"},
    {"slug": "pend_etiqueta","nome": "Pendências de etiqueta",          "icone": "🏷", "liberado": True, "endpoint": "almox.pendencias_etiqueta"},
    {"slug": "colaboradores","nome": "Colaboradores",                  "icone": "👤", "liberado": True, "endpoint": "almox.colaboradores", "somente_colab": True},
    {"slug": "papeis",      "nome": "Papéis de colaborador",           "icone": "🎭", "liberado": True, "endpoint": "almox.papeis", "somente_admin": True},
]


def modulo_required(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not _efetivo("pode_almox_modulo"):
            if _ver_como_nome():
                return _bloqueado_ver_como()
            abort(403)
        return f(*args, **kwargs)
    return wrapper


def _guard(prop):
    def deco(f):
        @wraps(f)
        @login_required
        def wrapper(*args, **kwargs):
            if not _efetivo(prop):
                if _ver_como_nome():
                    return _bloqueado_ver_como()
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return deco


def _log(categoria, detalhe):
    from .models import Usuario as _U
    autor_id = current_user.id if isinstance(current_user, _U) else None
    db.session.add(AlmoxLog(autor_id=autor_id, autor_nome=getattr(current_user, "nome", None),
                            categoria=categoria, detalhe=detalhe))


def _venc_info(validade):
    """Retorna (rotulo, cor_bootstrap) a partir da data de validade."""
    if not validade:
        return ("—", "secondary")
    dias = (validade - date.today()).days
    if dias < 0:
        return ("Vencido", "danger")
    if dias <= 30:
        return (f"Vence em {dias}d", "danger")
    if dias <= 90:
        return (f"Vence em {dias}d", "warning")
    return ("No prazo", "success")


@almox_bp.route("/")
@modulo_required
def home():
    cap = {
        "coletor": _efetivo("pode_coletor"),
        "material": _efetivo("pode_material"),
        "chaves": _efetivo("pode_chaves"),
        "extintores": _efetivo("pode_extintores"),
        "solic": _efetivo("pode_ver_solicitacoes"),
        "relatorios": _efetivo("pode_relatorios"),
        "colaboradores": _efetivo("pode_colaboradores"),
    }
    is_admin = _efetivo("is_admin")

    # Ações rápidas disponíveis (rótulo, ícone-emoji, endpoint) conforme permissão
    ACOES = []
    if cap["coletor"]:
        ACOES.append(("Abrir coletor", "📷", "almox.coletor"))
    if cap["solic"]:
        ACOES.append(("Minhas solicitações", "📝", "solicitante.index"))
    if cap["material"]:
        ACOES.append(("Material (estoque)", "📦", "almox.materiais"))
        ACOES.append(("Confirmar chegadas", "📥", "almox.index"))
    if cap["chaves"]:
        ACOES.append(("Chaves", "🔑", "almox.chaves"))
    if cap["extintores"]:
        ACOES.append(("Extintores", "🧯", "almox.extintores"))

    # "amplo" -> painel; senão -> quiosque na ação principal
    areas = [k for k in ("coletor", "material", "chaves", "extintores", "solic") if cap[k]]
    amplo = is_admin or cap["material"] or cap["colaboradores"] or len(areas) >= 2
    modo = "painel" if amplo else "quiosque"
    principal = None
    if modo == "quiosque":
        ordem = ["coletor", "solic", "chaves", "extintores", "material"]
        alvo = next((a for a in ordem if cap[a]), None)
        _mapa = {
            "coletor": ("Abrir Coletor", "📷", "almox.coletor", "Retirar e devolver pelo QR"),
            "solic": ("Minhas solicitações", "📝", "solicitante.index", "Acompanhe e abra novas"),
            "chaves": ("Chaves", "🔑", "almox.chaves", "Retirar e devolver chaves"),
            "extintores": ("Extintores", "🧯", "almox.extintores", "Inspeções e reposição"),
            "material": ("Material", "📦", "almox.materiais", "Estoque e movimentações"),
        }
        principal = _mapa.get(alvo)

    # Pendências (cada uma gated pela permissão de quem pode agir) — com link de atalho
    pend = []
    tiles = []
    listas_pend = []
    try:
        if is_admin:
            nb = AjusteInventario.query.filter_by(status="pendente").count()
            if nb:
                pend.append((nb, "Baixas de inventário a aprovar", url_for("almox.inventario_pendentes")))
            nnf = NotaFiscalAlmox.query.filter(NotaFiscalAlmox.classificacao.is_(None)).count()
            if nnf:
                pend.append((nnf, "NFs a classificar (OPEX/CAPEX)", url_for("almox.administrativo")))
        if cap["extintores"]:
            cont = {"IRREGULAR": 0, "VENCIDO": 0, "PROX_VENC": 0, "EM_RECARGA": 0, "PRONTO_REPO": 0, "ATENCAO": 0}
            for e in Extintor.query.filter_by(ativo=True).all():
                k = _situacao_extintor(e)[0]
                if k in cont:
                    cont[k] += 1
            _purl = url_for("almox.pendencias_etiqueta")
            irr = cont["IRREGULAR"] + cont["VENCIDO"]
            if irr:
                pend.append((irr, "Extintores irregulares / vencidos", _purl))
            if cont["PROX_VENC"]:
                pend.append((cont["PROX_VENC"], "Extintores próximos do vencimento", _purl))
            outros = cont["EM_RECARGA"] + cont["PRONTO_REPO"] + cont["ATENCAO"]
            if outros:
                pend.append((outros, "Em recarga / reposição / atenção", _purl))
            n_etq = PendenciaEtiqueta.query.filter_by(resolvida=False).count()
            if n_etq:
                pend.append((n_etq, "Pendências de etiqueta", _purl))
        if cap["material"]:
            mb = sum(1 for p in ProdutoAlmox.query.filter_by(ativo=True).all() if p.abaixo_minimo)
            if mb:
                tiles.append((mb, "Abaixo do mínimo", "warn", url_for("almox.materiais", baixo="1")))
        if cap["chaves"]:
            eu = Chave.query.filter_by(ativo=True, status="Em uso").count()
            tiles.append((eu, "Chaves em uso", "", url_for("almox.chaves")))
    except Exception:
        pass

    return render_template("almox/home.html", modo=modo, principal=principal,
                           acoes=ACOES, pend=pend, tiles=tiles, listas_pend=listas_pend, cap=cap)


@almox_bp.route("/em-construcao/<slug>")
@modulo_required
def em_construcao(slug):
    topico = next((t for t in TOPICOS if t["slug"] == slug), None)
    if not topico:
        abort(404)
    return render_template("almox/em_construcao.html", topico=topico)


# ---------- CHAVES ----------
@almox_bp.route("/chaves")
@_guard("pode_chaves")
def chaves():
    q = (request.args.get("q") or "").strip()
    inativos = request.args.get("inativos") == "1"
    quadros_sel = _args_list("quadro")
    status_sel = _args_list("status")
    base = Chave.query.filter_by(ativo=False) if inativos else Chave.query.filter_by(ativo=True)
    itens = []
    for c in base.order_by(Chave.descricao).all():
        if q and not contem_busca(" ".join([c.descricao or "", c.quadro_nome or "", c.status or "", c.com_quem or ""]), q):
            continue
        if quadros_sel and (c.quadro_nome or "") not in quadros_sel:
            continue
        if status_sel and (c.status or "") not in status_sel:
            continue
        itens.append(c)
    quadros = QuadroChave.query.filter_by(ativo=True).order_by(QuadroChave.nome).all()
    colabs = Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome).all()
    n_inativas = Chave.query.filter_by(ativo=False).count()
    atrasadas = [l for l in _chaves_situacao() if l["atrasada"]]   # 48b: alerta ao entrar em chaves
    return render_template("almox/chaves.html", itens=itens, q=q, quadros=quadros, colabs=colabs,
                           inativos=inativos, n_inativas=n_inativas,
                           quadros_sel=quadros_sel, status_sel=status_sel, atrasadas=atrasadas)


def _resolver_quadro(form):
    """Resolve o quadro pelo id (select) ou, como fallback, pelo nome digitado."""
    qid = form.get("quadro_chave_id") or None
    if qid and str(qid).isdigit():
        return int(qid)
    nome = (form.get("quadro_nome") or form.get("quadro") or "").strip()
    if nome:
        q = QuadroChave.query.filter(db.func.upper(QuadroChave.nome) == nome.upper()).first()
        if q:
            return q.id
    return None


@almox_bp.route("/chaves/nova", methods=["POST"])
@_guard("pode_chaves")
def chave_nova():
    import secrets
    desc = (request.form.get("descricao") or "").strip()
    if not desc:
        flash("Informe a descrição da chave.", "danger")
        return redirect(url_for("almox.chaves"))
    quadro_id = _resolver_quadro(request.form)
    uid = "CH-" + secrets.token_hex(4).upper()
    while Chave.query.filter_by(qr_uid=uid).first():
        uid = "CH-" + secrets.token_hex(4).upper()
    c = Chave(descricao=desc.upper(), quadro_chave_id=quadro_id,
              qr_uid=uid, status="Disponível")
    db.session.add(c)
    _log("Chave", f"Chave cadastrada: {c.descricao}" + (f" (quadro: {c.quadro_nome})" if quadro_id else " (sem quadro)"))
    db.session.commit()
    flash("Chave cadastrada." + ("" if quadro_id else " (Sem quadro — você pode definir depois em Editar.)"), "success")
    return redirect(url_for("almox.chaves"))


@almox_bp.route("/chaves/<int:cid>/desativar", methods=["POST"])
@_guard("chave_desativar")
def chave_desativar(cid):
    c = db.session.get(Chave, cid) or abort(404)
    if c.status == "Em uso":
        flash("Chave em uso não pode ser desativada. Faça a devolução primeiro.", "warning")
        return redirect(url_for("almox.chaves"))
    c.ativo = False
    _log("Chave", f"Chave {c.descricao} desativada por {current_user.nome}")
    db.session.commit()
    flash("Chave desativada.", "success")
    return redirect(url_for("almox.chaves"))


@almox_bp.route("/chaves/<int:cid>/reativar", methods=["POST"])
@_guard("chave_desativar")
def chave_reativar(cid):
    c = db.session.get(Chave, cid) or abort(404)
    c.ativo = True
    _log("Chave", f"Chave {c.descricao} reativada por {current_user.nome}")
    db.session.commit()
    flash("Chave reativada.", "success")
    return redirect(url_for("almox.chaves", inativos="1"))


@almox_bp.route("/chaves/<int:cid>/editar", methods=["POST"])
@_guard("pode_chaves")
def chave_editar(cid):
    c = db.session.get(Chave, cid) or abort(404)
    mudancas = []
    nova_desc = (request.form.get("descricao") or "").strip().upper()
    if nova_desc and nova_desc != (c.descricao or ""):
        mudancas.append(f"descrição: {c.descricao or '—'} → {nova_desc}")
        c.descricao = nova_desc
    novo_quadro = _resolver_quadro(request.form)
    if novo_quadro != c.quadro_chave_id:
        antigo = c.quadro_nome
        c.quadro_chave_id = novo_quadro
        mudancas.append(f"quadro: {antigo or '—'} → {c.quadro_nome or '—'}")
    if mudancas:
        _log("Chave", f"Chave {c.descricao} editada — " + "; ".join(mudancas))
        db.session.commit()
        flash("Chave atualizada.", "success")
    else:
        flash("Nada a alterar.", "info")
    return redirect(url_for("almox.chaves"))


@almox_bp.route("/chaves/<int:cid>/historico")
@_guard("pode_chaves")
def chave_historico(cid):
    c = db.session.get(Chave, cid) or abort(404)
    movs = (MovimentacaoChave.query.filter_by(chave_id=c.id)
            .order_by(MovimentacaoChave.criado_em.desc()).all())
    return render_template("almox/chave_historico.html", c=c, movs=movs)


def _op_id():
    """id do operador para gravar em operador_id (FK usuarios). None se for colaborador."""
    from .models import Usuario as _U
    return current_user.id if isinstance(current_user, _U) else None


def _resolver_colaborador(termo):
    """Acha o colaborador por QR (COL-...), CPF (só dígitos) ou nome. Devolve Colaborador ou None."""
    t = (termo or "").strip()
    if not t:
        return None
    # QR: 'COLAB:COL-XXXX' ou 'COL-XXXX'
    uid = t.split(":", 1)[1] if ":" in t else t
    c = Colaborador.query.filter_by(qr_uid=uid, ativo=True).first()
    if c:
        return c
    # CPF (só dígitos)
    dig = "".join(ch for ch in t if ch.isdigit())
    if dig:
        for col in Colaborador.query.filter(Colaborador.ativo.is_(True)).all():
            if "".join(ch for ch in (col.cpf or "") if ch.isdigit()) == dig:
                return col
    # Nome (exato, ignorando caixa)
    return Colaborador.query.filter(db.func.upper(Colaborador.nome) == t.upper(),
                                    Colaborador.ativo.is_(True)).first()


@almox_bp.route("/chaves/<int:cid>/toggle", methods=["POST"])
@_guard("pode_chaves")
def chave_toggle(cid):
    c = db.session.get(Chave, cid) or abort(404)
    if c.status == "Disponível":
        termo = request.form.get("com_quem") or ""
        colab = _resolver_colaborador(termo)
        if not colab:
            flash("Informe quem vai retirar (nome, CPF ou bipe o QR de um colaborador cadastrado).", "danger")
            return redirect(url_for("almox.chaves"))
        # Retirada pede a senha do colaborador
        senha = request.form.get("senha") or ""
        if not colab.tem_senha:
            flash(f"{colab.nome} ainda não definiu a senha. Ele define no 1º acesso (login por CPF) ou peça reset.", "warning")
            return redirect(url_for("almox.chaves"))
        if not colab.check_senha(senha):
            flash("Senha do colaborador inválida. Retirada não confirmada.", "danger")
            return redirect(url_for("almox.chaves"))
        c.status = "Em uso"
        c.com_quem = colab.nome
        db.session.add(MovimentacaoChave(
            chave_id=c.id, chave_desc=c.descricao, quadro_nome=c.quadro_nome,
            colaborador_id=colab.id, colaborador_nome=colab.nome,
            acao="retirada", operador_id=_op_id()))
        _log("Chave", f"Chave {c.descricao} retirada por {colab.nome}")
    else:
        retirou = c.com_quem or "—"
        termo = request.form.get("devolvido_por") or ""
        colab = _resolver_colaborador(termo)
        devolvedor = colab.nome if colab else (termo.strip().upper() or retirou)
        db.session.add(MovimentacaoChave(
            chave_id=c.id, chave_desc=c.descricao, quadro_nome=c.quadro_nome,
            colaborador_id=(colab.id if colab else None), colaborador_nome=devolvedor,
            retirado_por=retirou, acao="devolucao",
            operador_id=_op_id()))
        _log("Chave", f"Chave {c.descricao} devolvida por {devolvedor}" +
             (f" (estava com {retirou})" if devolvedor != retirou else ""))
        c.status = "Disponível"
        c.com_quem = None
    db.session.commit()
    return redirect(url_for("almox.chaves"))


# ----- Quadros de Chave (localizador) -----
def _backfill_qr_quadros():
    """Garante que todo quadro tenha um QR próprio (QUAD-...)."""
    import secrets
    faltando = QuadroChave.query.filter((QuadroChave.qr_uid.is_(None)) | (QuadroChave.qr_uid == "")).all()
    for q in faltando:
        uid = "QUAD-" + secrets.token_hex(4).upper()
        while QuadroChave.query.filter_by(qr_uid=uid).first():
            uid = "QUAD-" + secrets.token_hex(4).upper()
        q.qr_uid = uid
    if faltando:
        db.session.commit()


@almox_bp.route("/chaves/quadros")
@_guard("pode_chaves")
def quadros_chave():
    _backfill_qr_quadros()
    quadros = QuadroChave.query.order_by(QuadroChave.nome).all()
    return render_template("almox/quadros_chave.html", quadros=quadros)


@almox_bp.route("/chaves/quadros/novo", methods=["POST"])
@_guard("pode_chaves")
def quadro_novo():
    import secrets
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        flash("Informe o nome do Quadro de Chaves.", "danger")
        return redirect(url_for("almox.quadros_chave"))
    if QuadroChave.query.filter(db.func.upper(QuadroChave.nome) == nome.upper()).first():
        flash("Já existe um Quadro de Chaves com esse nome.", "warning")
        return redirect(url_for("almox.quadros_chave"))
    uid = "QUAD-" + secrets.token_hex(4).upper()
    while QuadroChave.query.filter_by(qr_uid=uid).first():
        uid = "QUAD-" + secrets.token_hex(4).upper()
    db.session.add(QuadroChave(nome=nome.upper(), qr_uid=uid))
    _log("Chave", f"Quadro de Chaves cadastrado: {nome.upper()}")
    db.session.commit()
    flash("Quadro de Chaves cadastrado.", "success")
    return redirect(url_for("almox.quadros_chave"))


@almox_bp.route("/chaves/quadros/<int:qid>/toggle", methods=["POST"])
@_guard("pode_chaves")
def quadro_toggle(qid):
    quadro = db.session.get(QuadroChave, qid) or abort(404)
    quadro.ativo = not quadro.ativo
    db.session.commit()
    return redirect(url_for("almox.quadros_chave"))


# ----- QR individual da chave (impressão) -----
@almox_bp.route("/chaves/qr")
@_guard("pode_chaves")
def chaves_qr():
    """Página de impressão dos QR das chaves (uma etiqueta por chave)."""
    ids = request.args.get("ids") or ""
    consulta = Chave.query.filter_by(ativo=True)
    if ids.strip():
        try:
            lista_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
            consulta = consulta.filter(Chave.id.in_(lista_ids))
        except ValueError:
            pass
    itens = consulta.order_by(Chave.descricao).all()
    formato = request.args.get("formato", "a4")   # 'a4' | 'termica' | 'dobravel'
    return render_template("almox/chaves_qr.html", itens=itens, qr_svg=_qr_svg, formato=formato)


@almox_bp.route("/chaves/quadros/qr")
@_guard("pode_chaves")
def quadros_qr():
    """Página de impressão dos QR dos quadros (uma etiqueta por quadro)."""
    _backfill_qr_quadros()
    ids = request.args.get("ids") or ""
    consulta = QuadroChave.query.filter_by(ativo=True)
    if ids.strip():
        lista_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        if lista_ids:
            consulta = consulta.filter(QuadroChave.id.in_(lista_ids))
    itens = consulta.order_by(QuadroChave.nome).all()
    formato = request.args.get("formato", "a4")
    return render_template("almox/quadros_qr.html", itens=itens, qr_svg=_qr_svg, formato=formato)

# ---------- EXTINTORES (Etapa 3) ----------
import json as _json
from datetime import datetime as _dt

MESES_PT = ["", "JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

SITUACAO_LABEL = {
    "NO_PRAZO": ("No prazo", "success"),
    "PROX_VENC": ("Próximo do vencimento", "warning"),
    "VENCIDO": ("Irregular / Vencido", "danger"),
    "IRREGULAR": ("Irregular / Vencido", "danger"),
    "ATENCAO": ("Atenção (etiqueta)", "info"),
    "EM_RECARGA": ("Em recarga", "info"),
    "PRONTO_REPO": ("Pronto p/ reposição", "primary"),
}

# Opções do filtro (sem duplicar "Irregular / Vencido": VENCIDO cobre também IRREGULAR)
SITUACAO_OPCOES = [
    ("NO_PRAZO", "No prazo"),
    ("PROX_VENC", "Próximo do vencimento"),
    ("VENCIDO", "Irregular / Vencido"),
    ("ATENCAO", "Atenção (etiqueta)"),
    ("EM_RECARGA", "Em recarga"),
    ("PRONTO_REPO", "Pronto p/ reposição"),
]


def _match_situacao(sel, k):
    """Filtro de situação: 'VENCIDO' cobre tanto o vencido por data quanto o irregular operacional."""
    if not sel:
        return True
    if sel == "VENCIDO":
        return k in ("VENCIDO", "IRREGULAR")
    return k == sel


def _args_list(nome):
    """Valores marcados de um filtro multi-selecao (getlist), sem vazios."""
    return [x for x in request.args.getlist(nome) if x]


def _situacoes_sel():
    """Lista de situacoes marcadas (multi-selecao). Vazia = todas."""
    return [x for x in request.args.getlist("situacao") if x]


def _match_situacoes(sels, k):
    return True if not sels else any(_match_situacao(x, k) for x in sels)


def _competencia(d):
    """Rótulo MMM/AAAA de uma data (competência)."""
    return f"{MESES_PT[d.month]}/{d.year}" if d else "—"


def _meses_ate(d):
    if not d:
        return None
    hoje = date.today()
    return (d.year - hoje.year) * 12 + (d.month - hoje.month)


def _situacao_extintor(e):
    """Ciclo do extintor. Estados operacionais gravados em e.situacao
    (IRREGULAR/ATENCAO/EM_RECARGA/PRONTO_REPO); NO_PRAZO deriva PROX/VENCIDO das datas."""
    s = e.situacao or "NO_PRAZO"
    if s in ("EM_RECARGA", "PRONTO_REPO", "IRREGULAR", "ATENCAO"):
        return (s,) + SITUACAO_LABEL[s]
    piores = []
    for d in (e.validade, e.teste_hidrostatico):
        m = _meses_ate(d)
        if m is None:
            continue
        if m < 0:
            piores.append("VENCIDO")
        elif m <= 1:
            piores.append("PROX_VENC")
        else:
            piores.append("NO_PRAZO")
    if "VENCIDO" in piores:
        k = "VENCIDO"
    elif "PROX_VENC" in piores:
        k = "PROX_VENC"
    else:
        k = "NO_PRAZO"
    return (k,) + SITUACAO_LABEL[k]


def _parse_mmaaaa(prefixo, form):
    """Lê dois selects (mes_<prefixo>, ano_<prefixo>) e devolve date no dia 01, ou None."""
    return _parse_mmaaaa_campos(f"mes_{prefixo}", f"ano_{prefixo}", form)


def _parse_mmaaaa_campos(mes_field, ano_field, form):
    mes = form.get(mes_field) or ""
    ano = form.get(ano_field) or ""
    if mes.isdigit() and ano.isdigit():
        try:
            return date(int(ano), int(mes), 1)
        except ValueError:
            return None
    return None


def _parse_ano(prefixo, form):
    """Teste hidrostático é ANUAL: lê só ano_<prefixo> e devolve 31/12 daquele ano (vale o ano todo)."""
    return _parse_ano_campo(f"ano_{prefixo}", form)


def _parse_ano_campo(ano_field, form):
    ano = form.get(ano_field) or ""
    if ano.isdigit():
        try:
            return date(int(ano), 12, 31)
        except ValueError:
            return None
    return None


def _th_label(d):
    """Rótulo do teste hidrostático (só o ano)."""
    return str(d.year) if d else "—"


def _anos_range():
    a = date.today().year
    return list(range(a - 6, a + 8))


def _colab_sessao():
    """Colaborador autenticado no fluxo de campo (via QR), guardado na sessão."""
    cid = session.get("colab_ext_id")
    return db.session.get(Colaborador, cid) if cid else None


def _pode_gerir_ext():
    """True se o ator atual pode fazer as ações de gestão (regularizar/conferir/repor)."""
    return bool(current_user.is_authenticated and getattr(current_user, "pode_extintores", False))


def _ext_acesso(f):
    """Libera a rota para: usuário logado com acesso a extintores OU colaborador de campo (sessão)."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if current_user.is_authenticated and getattr(current_user, "pode_extintores", False):
            return f(*args, **kwargs)
        if _colab_sessao():
            return f(*args, **kwargs)
        eid = kwargs.get("eid")
        e = db.session.get(Extintor, eid) if eid else None
        if e and e.qr_uid:
            return redirect(url_for("almox.extintor_campo", qr_uid=e.qr_uid))
        return redirect(url_for("auth.login"))
    return wrapper


def _redir_ficha(e):
    """Volta para a ficha certa conforme o ator (desktop logado x campo/colaborador)."""
    if current_user.is_authenticated:
        return redirect(url_for("almox.extintor_ficha", eid=e.id))
    return redirect(url_for("almox.extintor_campo", qr_uid=e.qr_uid))



def _dias_uteis(inicio, fim):
    """Conta dias UTEIS (seg-sex) entre inicio e fim."""
    from datetime import timedelta
    if not inicio:
        return 0
    d = inicio.date() if hasattr(inicio, "date") else inicio
    f = fim.date() if hasattr(fim, "date") else fim
    dias = 0
    while d < f:
        d += timedelta(days=1)
        if d.weekday() < 5:
            dias += 1
    return dias


LIMITE_DIAS_CHAVE = 3   # alerta quando colaborador esta com a chave ha mais de X dias uteis


def _chaves_situacao():
    """Situacao atual de cada chave ativa: no quadro ou com colaborador (desde quando, dias uteis)."""
    linhas = []
    agora = datetime.utcnow()
    for c in Chave.query.filter_by(ativo=True).order_by(Chave.descricao).all():
        em_uso = (c.status or "").strip().lower() == "em uso"
        desde = None
        dias = 0
        if em_uso:
            mv = (MovimentacaoChave.query.filter_by(chave_id=c.id, acao="retirada")
                  .order_by(MovimentacaoChave.criado_em.desc()).first())
            desde = mv.criado_em if mv else None
            dias = _dias_uteis(desde, agora) if desde else 0
        linhas.append({"chave": c, "em_uso": em_uso, "com_quem": c.com_quem or "—",
                       "quadro_nome": _quadro_nome_chave(c),
                       "desde": desde, "dias": dias,
                       "atrasada": bool(em_uso and dias > LIMITE_DIAS_CHAVE)})
    return linhas


def _quadro_nome_chave(c):
    try:
        q = getattr(c, "quadro", None) or getattr(c, "quadro_chave", None)
        if q is not None and getattr(q, "nome", None):
            return q.nome
    except Exception:
        pass
    return c.local or "—"


@almox_bp.route("/chaves/situacao")
@_guard("pode_chaves")
def relatorio_chaves_situacao():
    linhas = _chaves_situacao()
    atrasadas = [l for l in linhas if l["atrasada"]]
    return render_template("almox/chaves_situacao.html", linhas=linhas,
                           atrasadas=atrasadas, limite=LIMITE_DIAS_CHAVE)


@almox_bp.route("/extintores")
@_guard("pode_extintores")
def extintores():
    from .seed_extintores import PREDIO_LABEL
    predios_sel = _args_list("predio")
    locais_sel = _args_list("local")
    tipos_sel = _args_list("tipo")
    situacoes = _situacoes_sel()
    linhas = []
    todos = Extintor.query.filter_by(ativo=True).order_by(Extintor.predio, Extintor.local, Extintor.codigo).all()
    for e in todos:
        if predios_sel and e.predio not in predios_sel:
            continue
        if tipos_sel and e.tipo not in tipos_sel:
            continue
        if locais_sel and e.local not in locais_sel:
            continue
        k, lbl, cls = _situacao_extintor(e)
        if not _match_situacoes(situacoes, k):
            continue
        linhas.append((e, k, lbl, cls))
    predios = sorted({e.predio for e in todos if e.predio})
    tipos = sorted({e.tipo for e in todos if e.tipo})
    # Locais disponiveis conforme os OUTROS filtros (predio/tipo/situacao) — dependente
    locais = sorted({e.local for e in todos
                     if e.local
                     and (not predios_sel or e.predio in predios_sel)
                     and (not tipos_sel or e.tipo in tipos_sel)
                     and _match_situacoes(situacoes, _situacao_extintor(e)[0])})
    n_pend = PendenciaEtiqueta.query.filter_by(resolvida=False).count()
    return render_template("almox/extintores.html", linhas=linhas,
                           predios_sel=predios_sel, locais_sel=locais_sel, tipos_sel=tipos_sel,
                           situacoes=situacoes, predios=predios, tipos=tipos, locais=locais,
                           predio_label=PREDIO_LABEL, situacao_label=SITUACAO_LABEL,
                           situacao_opcoes=SITUACAO_OPCOES,
                           competencia=_competencia, n_pend=n_pend)


@almox_bp.route("/extintores/<int:eid>")
@_ext_acesso
def extintor_ficha(eid):
    import json as _json
    e = db.session.get(Extintor, eid) or abort(404)
    k, lbl, cls = _situacao_extintor(e)
    hist_raw = (InspecaoExtintor.query.filter_by(extintor_id=e.id)
                .order_by(InspecaoExtintor.criado_em.desc()).limit(30).all())
    hist = []
    for h in hist_raw:
        try:
            itens = _json.loads(h.itens_json) if h.itens_json else {}
        except Exception:
            itens = {}
        hist.append({"h": h, "itens": itens})
    colab = _colab_sessao() if not current_user.is_authenticated else None
    # itens do checklist de retorno (almox) = sem "Acesso e sinalização"
    check_retorno = [c for c in CHECK_EXTINTOR if not c.lower().startswith("acesso")]
    return render_template("almox/extintor_ficha.html", e=e, sit_k=k, sit_lbl=lbl, sit_cls=cls,
                           check=CHECK_EXTINTOR, check_retorno=check_retorno,
                           item_etiqueta=ITEM_ETIQUETA_EXTINTOR,
                           hist=hist, competencia=_competencia, meses=MESES_PT,
                           anos=_anos_range(), pode_gerir=_pode_gerir_ext(),
                           campo=(not current_user.is_authenticated), ator_colab=colab)


def _coletar_checklist(form):
    """Lê o checklist. Devolve (itens_dict, core_nok, etiqueta_nok).
    O item da etiqueta é separado (regra especial de 'Atenção')."""
    itens = {}
    core_nok = False
    for i, item in enumerate(CHECK_EXTINTOR):
        v = form.get(f"item_{i}")
        if v is None:
            continue                      # item não exibido nesse checklist (ex.: retorno)
        itens[item] = v
        if v == "nok":
            core_nok = True
    et = form.get("item_etiqueta", "na")
    itens[ITEM_ETIQUETA_EXTINTOR] = et
    return itens, core_nok, (et == "nok")


def _quem(form):
    """Quem operou: colaborador de campo (sessão) ou o usuário logado. Não usa mais campo de texto."""
    colab = _colab_sessao()
    if colab and not current_user.is_authenticated:
        return colab.nome, colab.id
    if current_user.is_authenticated:
        return current_user.nome, None
    return ("—", None)


def _aplicar_resultado(e, core_nok, et_nok, nome):
    """Define a situação do extintor conforme o checklist e trata pendências."""
    if core_nok:
        e.situacao = "IRREGULAR"          # regularização pendente até voltar ao local
        return "irregular"
    if et_nok:
        e.situacao = "ATENCAO"            # segue em uso; só pendência de etiqueta
        if not PendenciaEtiqueta.query.filter_by(extintor_id=e.id, resolvida=False).first():
            db.session.add(PendenciaEtiqueta(extintor_id=e.id, extintor_cod=e.codigo,
                           predio=e.predio, local=e.local, aberta_por=nome))
        return "irregular"
    e.situacao = "NO_PRAZO"
    return "conforme"


@almox_bp.route("/extintores/<int:eid>/inspecionar", methods=["POST"])
@_ext_acesso
def extintor_inspecionar(eid):
    import json as _json
    e = db.session.get(Extintor, eid) or abort(404)
    itens, core_nok, et_nok = _coletar_checklist(request.form)
    # Item especial: datas de recarga/TH conferem com o extintor? Se "nao", corrige as datas.
    datas_conf = request.form.get("item_datas")  # "sim" | "nao" | None
    if datas_conf == "nao":
        nv = _parse_mmaaaa_campos("corr_mes_validade", "corr_ano_validade", request.form)
        nth = _parse_ano_campo("corr_ano_th", request.form)
        corr = []
        if nv:
            e.validade = nv
            corr.append(f"validade -> {_competencia(nv)}")
        if nth:
            e.teste_hidrostatico = nth
            corr.append(f"TH -> {_th_label(nth)}")
        itens["Datas conferem com o app?"] = "não (corrigidas: " + ", ".join(corr) + ")" if corr else "não"
        if corr:
            _log("Extintor", f"{e.codigo}: datas ajustadas na inspeção ({'; '.join(corr)})")
    elif datas_conf:
        itens["Datas conferem com o app?"] = datas_conf
    nova_classe = (request.form.get("corr_classe") or "").strip().upper()
    if nova_classe and nova_classe != (e.classe or "").upper():
        antiga = e.classe or "—"
        e.classe = nova_classe
        itens["Classe corrigida"] = f"{antiga} → {nova_classe}"
        _log("Extintor", f"{e.codigo}: classe alterada de {antiga} para {nova_classe} na inspeção")
    nome, cid = _quem(request.form)
    resultado = _aplicar_resultado(e, core_nok, et_nok, nome)
    e.inspecao = date.today()
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="inspecao",
                   resultado=resultado, itens_json=_json.dumps(itens, ensure_ascii=False),
                   etiqueta_ok=(not et_nok), obs=(request.form.get("obs") or "").strip(),
                   colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None)))
    if core_nok:
        _log("Extintor", f"{e.codigo} ({e.local}): inspeção IRREGULAR por {nome} — notificar ADMIN")
        msg, cat = "Inspeção registrada. Extintor IRREGULAR — leve ao Almox D6.", "warning"
    elif et_nok:
        _log("Extintor", f"{e.codigo}: etiqueta em desacordo por {nome} — pendência aberta (Atenção)")
        msg, cat = "Inspeção registrada. Etiqueta em desacordo: status Atenção + pendência.", "warning"
    else:
        _log("Extintor", f"{e.codigo} ({e.local}): inspeção conforme por {nome}")
        msg, cat = "Inspeção conforme registrada.", "success"
    db.session.commit()
    flash(msg, cat)
    return _redir_ficha(e)


@almox_bp.route("/extintores/<int:eid>/reposicao", methods=["POST"])
@_ext_acesso
def extintor_reposicao(eid):
    """Troca programada: substitui o extintor, faz o checklist do NOVO e lança nova validade/TH."""
    import json as _json
    e = db.session.get(Extintor, eid) or abort(404)
    itens, core_nok, et_nok = _coletar_checklist(request.form)
    nome, cid = _quem(request.form)
    nova_val = _parse_mmaaaa("validade", request.form)
    novo_th = _parse_ano("th", request.form)
    if nova_val:
        e.validade = nova_val
    if novo_th:
        e.teste_hidrostatico = novo_th
    resultado = _aplicar_resultado(e, core_nok, et_nok, nome)
    e.inspecao = date.today()
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="reposicao",
                   resultado=resultado, itens_json=_json.dumps(itens, ensure_ascii=False),
                   etiqueta_ok=(not et_nok), colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None),
                   obs=f"Troca programada. Validade {_competencia(e.validade)}, TH {_th_label(e.teste_hidrostatico)}"))
    _log("Extintor", f"{e.codigo}: reposição/troca por {nome} "
                     f"(validade {_competencia(e.validade)}, TH {_th_label(e.teste_hidrostatico)})")
    db.session.commit()
    flash("Reposição (troca) registrada.", "success")
    return _redir_ficha(e)


@almox_bp.route("/extintores/<int:eid>/regularizar", methods=["POST"])
@_ext_acesso
def extintor_regularizar(eid):
    """Único caminho na ficha Irregular: Levado ao Almox D6 (sem pedir nome)."""
    e = db.session.get(Extintor, eid) or abort(404)
    nome, cid = _quem(request.form)
    e.situacao = "EM_RECARGA"
    e.retirado_por = nome
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="retirada",
                   resultado="irregular", colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None)))
    _log("Extintor", f"{e.codigo}: levado ao Almox D6 p/ recarga por {nome}")
    db.session.commit()
    flash("Extintor marcado como Em recarga (levado ao Almox D6).", "info")
    return _redir_ficha(e)


@almox_bp.route("/extintores/<int:eid>/conferir", methods=["POST"])
@_ext_acesso
def extintor_conferir(eid):
    """Conferência do Almoxarifado no retorno (sem 'Acesso', sem nome) → Pronto p/ reposição."""
    import json as _json
    e = db.session.get(Extintor, eid) or abort(404)
    if not _pode_gerir_ext():
        abort(403)
    itens, core_nok, et_nok = _coletar_checklist(request.form)
    nome, cid = _quem(request.form)
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="conferencia",
                   resultado=("irregular" if (core_nok or et_nok) else "conforme"),
                   itens_json=_json.dumps(itens, ensure_ascii=False), etiqueta_ok=(not et_nok),
                   colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None)))
    if et_nok and not PendenciaEtiqueta.query.filter_by(extintor_id=e.id, resolvida=False).first():
        db.session.add(PendenciaEtiqueta(extintor_id=e.id, extintor_cod=e.codigo,
                       predio=e.predio, local=e.local, aberta_por=nome))
    e.situacao = "PRONTO_REPO"
    _log("Extintor", f"{e.codigo}: conferido no Almox (pronto p/ reposição) por {nome}")
    db.session.commit()
    flash("Conferência registrada. Extintor Pronto para reposição.", "success")
    return _redir_ficha(e)


@almox_bp.route("/extintores/<int:eid>/repor", methods=["POST"])
@_ext_acesso
def extintor_repor(eid):
    """Reposição final: volta ao local (sem nome), atualiza validade/TH (MMM+AAAA) → No prazo."""
    e = db.session.get(Extintor, eid) or abort(404)
    if not _pode_gerir_ext():
        abort(403)
    nome, cid = _quem(request.form)
    nova_val = _parse_mmaaaa("validade", request.form)
    novo_th = _parse_ano("th", request.form)
    if nova_val:
        e.validade = nova_val
    if novo_th:
        e.teste_hidrostatico = novo_th
    e.situacao = "NO_PRAZO"        # voltou ao local: sai da pendência de regularização
    e.retirado_por = None
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="reposicao",
                   resultado="conforme", colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None),
                   obs=f"Reposto no local. Validade {_competencia(e.validade)}, TH {_th_label(e.teste_hidrostatico)}"))
    _log("Extintor", f"{e.codigo}: reposto no local por {nome}")
    db.session.commit()
    flash("Reposição concluída. Extintor No prazo.", "success")
    return _redir_ficha(e)


@almox_bp.route("/extintores/<int:eid>/desativar", methods=["POST"])
@_guard("ext_desativar")
def extintor_desativar(eid):
    e = db.session.get(Extintor, eid) or abort(404)
    e.ativo = False
    _log("Extintor", f"{e.codigo}: extintor desativado por {current_user.nome}")
    db.session.commit()
    flash("Extintor desativado.", "success")
    return redirect(url_for("almox.extintores"))


@almox_bp.route("/extintores/cadastro")
@_guard("ext_cadastrar")
def extintor_cadastro():
    def distintos(col):
        vals = db.session.query(col).filter(col.isnot(None), col != "").distinct().all()
        return sorted({v[0] for v in vals if v[0]})
    sugestoes = {
        "predio": distintos(Extintor.predio),
        "local": distintos(Extintor.local),
        "tipo": distintos(Extintor.tipo),
        "classe": distintos(Extintor.classe),
    }
    return render_template("almox/extintor_cadastro.html", sugestoes=sugestoes,
                           check=CHECK_EXTINTOR, item_etiqueta=ITEM_ETIQUETA_EXTINTOR,
                           meses=MESES_PT, anos=_anos_range())


@almox_bp.route("/extintores/novo", methods=["POST"])
@_guard("ext_cadastrar")
def extintor_novo():
    import secrets, json as _json
    predio = (request.form.get("predio") or "").strip().upper()
    local = (request.form.get("local") or "").strip()
    tipo = (request.form.get("tipo") or "").strip().upper()
    classe = (request.form.get("classe") or "").strip().upper()
    if not local:
        flash("Informe ao menos o local do extintor.", "danger")
        return redirect(url_for("almox.extintores"))
    seq = (Extintor.query.count() or 0) + 1
    codigo = f"EXT{seq:04d}"
    while Extintor.query.filter_by(codigo=codigo).first():
        seq += 1
        codigo = f"EXT{seq:04d}"
    uid = "EXT-" + secrets.token_hex(4).upper()
    while Extintor.query.filter_by(qr_uid=uid).first():
        uid = "EXT-" + secrets.token_hex(4).upper()
    e = Extintor(codigo=codigo, predio=predio, local=local, tipo=tipo, classe=classe,
                 validade=_parse_mmaaaa("validade", request.form),
                 teste_hidrostatico=_parse_ano("th", request.form),
                 situacao="NO_PRAZO", status="No Local", qr_uid=uid, ativo=True)
    db.session.add(e)
    db.session.flush()
    # checklist inicial de conferência
    itens, core_nok, et_nok = _coletar_checklist(request.form)
    nome, cid = _quem(request.form)
    resultado = _aplicar_resultado(e, core_nok, et_nok, nome)
    e.inspecao = date.today()
    db.session.add(InspecaoExtintor(extintor_id=e.id, extintor_cod=e.codigo, tipo="inspecao",
                   resultado=resultado, itens_json=_json.dumps(itens, ensure_ascii=False),
                   etiqueta_ok=(not et_nok), colaborador_id=cid, colaborador_nome=nome,
                   operador_id=getattr(current_user, "id", None), obs="Cadastro / conferência inicial"))
    _log("Extintor", f"Extintor cadastrado: {codigo} ({local}) por {nome}")
    db.session.commit()
    flash(f"Extintor {codigo} cadastrado.", "success")
    return redirect(url_for("almox.extintor_ficha", eid=e.id))


# ----- Acesso pelo QR no campo (login único: CPF ou e-mail) -----
@almox_bp.route("/e/<qr_uid>")
def extintor_campo(qr_uid):
    e = Extintor.query.filter_by(qr_uid=qr_uid, ativo=True).first() or abort(404)
    # Já logado como usuário do sistema com acesso → ficha completa
    if current_user.is_authenticated and getattr(current_user, "pode_extintores", False):
        return _ficha_campo(e, pode_gerir=True, colab=None)
    # Colaborador de campo já autenticado nesta sessão → ficha de inspeção
    colab = _colab_sessao()
    if colab:
        return _ficha_campo(e, pode_gerir=False, colab=colab)
    # Senão, tela de login única
    return render_template("almox/extintor_login.html", e=e, etapa="login")


def _ficha_campo(e, pode_gerir, colab):
    import json as _json
    k, lbl, cls = _situacao_extintor(e)
    hist_raw = (InspecaoExtintor.query.filter_by(extintor_id=e.id)
                .order_by(InspecaoExtintor.criado_em.desc()).limit(20).all())
    hist = []
    for h in hist_raw:
        try:
            itens = _json.loads(h.itens_json) if h.itens_json else {}
        except Exception:
            itens = {}
        hist.append({"h": h, "itens": itens})
    check_retorno = [c for c in CHECK_EXTINTOR if not c.lower().startswith("acesso")]
    return render_template("almox/extintor_ficha.html", e=e, sit_k=k, sit_lbl=lbl, sit_cls=cls,
                           check=CHECK_EXTINTOR, item_etiqueta=ITEM_ETIQUETA_EXTINTOR,
                           check_retorno=check_retorno,
                           hist=hist, competencia=_competencia, meses=MESES_PT,
                           anos=_anos_range(), pode_gerir=pode_gerir, campo=True, ator_colab=colab)


@almox_bp.route("/e/<qr_uid>/entrar", methods=["POST"])
def extintor_campo_entrar(qr_uid):
    from .models import Usuario
    e = Extintor.query.filter_by(qr_uid=qr_uid, ativo=True).first() or abort(404)
    ident = (request.form.get("ident") or "").strip()
    senha = request.form.get("senha") or ""
    if "@" in ident:
        # login de usuário do sistema
        u = Usuario.query.filter_by(email=ident.lower()).first()
        if u and u.ativo and u.check_senha(senha):
            login_user(u)
            return redirect(url_for("almox.extintor_campo", qr_uid=qr_uid))
        flash("E-mail ou senha inválidos.", "danger")
        return render_template("almox/extintor_login.html", e=e, etapa="login")
    # senão, trata como CPF de colaborador
    cpf = "".join(c for c in ident if c.isdigit())
    colab = Colaborador.query.filter_by(ativo=True).filter(
        db.func.replace(db.func.replace(db.func.replace(Colaborador.cpf, ".", ""), "-", ""), " ", "") == cpf).first()
    if not colab:
        # fallback: compara só dígitos em Python (bases pequenas)
        colab = next((c for c in Colaborador.query.filter_by(ativo=True).all()
                      if "".join(ch for ch in (c.cpf or "") if ch.isdigit()) == cpf and cpf), None)
    if not colab:
        flash("CPF não encontrado no cadastro de colaboradores.", "danger")
        return render_template("almox/extintor_login.html", e=e, etapa="login")
    if not colab.tem_senha:
        # primeiro acesso: define a senha
        conf = request.form.get("confirma") or ""
        if not senha or len(senha) < 4:
            flash("Primeiro acesso: defina uma senha de ao menos 4 dígitos.", "warning")
            return render_template("almox/extintor_login.html", e=e, etapa="definir", colab_nome=colab.nome)
        if senha != conf:
            flash("As senhas não coincidem.", "danger")
            return render_template("almox/extintor_login.html", e=e, etapa="definir", colab_nome=colab.nome)
        colab.set_senha(senha)
        db.session.commit()
        session["colab_ext_id"] = colab.id
        flash("Senha definida. Bem-vindo!", "success")
        return redirect(url_for("almox.extintor_campo", qr_uid=qr_uid))
    if colab.check_senha(senha):
        session["colab_ext_id"] = colab.id
        return redirect(url_for("almox.extintor_campo", qr_uid=qr_uid))
    flash("Senha inválida.", "danger")
    return render_template("almox/extintor_login.html", e=e, etapa="login")


@almox_bp.route("/e/<qr_uid>/sair")
def extintor_campo_sair(qr_uid):
    session.pop("colab_ext_id", None)
    return redirect(url_for("almox.extintor_campo", qr_uid=qr_uid))


@almox_bp.route("/extintores/pendencias")
@_guard("pode_extintores")
def pendencias_etiqueta():
    abertas = (PendenciaEtiqueta.query.filter_by(resolvida=False)
               .order_by(PendenciaEtiqueta.aberta_em.desc()).all())
    resolvidas = (PendenciaEtiqueta.query.filter_by(resolvida=True)
                  .order_by(PendenciaEtiqueta.resolvida_em.desc()).limit(30).all())
    grupos = _pendencias_por_estado()
    return render_template("almox/pendencias_etiqueta.html", abertas=abertas,
                           resolvidas=resolvidas, grupos=grupos)


def _pendencias_por_estado():
    """Agrupa extintores pendentes (todos os estados != No prazo) por estado, para a pagina/PDF."""
    LABELS = [("IRREGULAR", "Irregulares / vencidos"), ("PROX_VENC", "Próximos do vencimento"),
              ("EM_RECARGA", "Em recarga"), ("PRONTO_REPO", "Prontos p/ reposição"),
              ("ATENCAO", "Atenção (etiqueta)")]
    buckets = {}
    for e in (Extintor.query.filter_by(ativo=True)
              .order_by(Extintor.predio, Extintor.local, Extintor.codigo).all()):
        k = _situacao_extintor(e)[0]
        if k == "NO_PRAZO":
            continue
        key = "IRREGULAR" if k in ("IRREGULAR", "VENCIDO") else k
        buckets.setdefault(key, []).append(e)
    return [{"chave": key, "titulo": lbl, "itens": buckets.get(key, [])}
            for key, lbl in LABELS if buckets.get(key)]


@almox_bp.route("/extintores/pendencias.pdf")
@_guard("pode_extintores")
def pendencias_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from flask import Response
    import io
    grupos = _pendencias_por_estado()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm, title="Pendências de extintores")
    st = getSampleStyleSheet()
    els = [Paragraph("Pendências de extintores — Serena · Cluster Delta", st["Title"]),
           Paragraph(datetime.now().strftime("Emitido em %d/%m/%Y %H:%M"), st["Normal"]),
           Spacer(1, 6 * mm)]
    total = 0
    for g in grupos:
        els.append(Paragraph(f"{g['titulo']} ({len(g['itens'])})", st["Heading3"]))
        dados = [["Código", "Prédio · Local", "Tipo/Carga · Classe", "Validade", "TH"]]
        for e in g["itens"]:
            total += 1
            loc = " · ".join([x for x in [e.predio, e.local] if x])
            val = _competencia(e.validade) if e.validade else "—"
            th = _th_label(e.teste_hidrostatico) if hasattr(e, "teste_hidrostatico") else "—"
            tc = (e.tipo or "—") + ((" · " + e.classe) if e.classe else "")
            dados.append([e.codigo or "—", loc, tc, val, th])
        t = Table(dados, colWidths=[26 * mm, 68 * mm, 46 * mm, 20 * mm, 15 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4B4B4B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F0EE")]),
        ]))
        els.append(t)
        els.append(Spacer(1, 5 * mm))
    if not total:
        els.append(Paragraph("Nenhuma pendência no momento.", st["Normal"]))
    doc.build(els)
    buf.seek(0)
    return Response(buf.read(), mimetype="application/pdf",
                    headers={"Content-Disposition": "inline; filename=pendencias_extintores.pdf"})


@almox_bp.route("/extintores/pendencias/<int:pid>/baixar", methods=["POST"])
@_guard("pode_extintores")
def pendencia_baixar(pid):
    p = db.session.get(PendenciaEtiqueta, pid) or abort(404)
    p.resolvida = True
    p.resolvida_em = _dt.utcnow()
    p.resolvida_por = current_user.nome
    # Se o extintor estava em "Atenção" só pela etiqueta, volta a No prazo
    ext = db.session.get(Extintor, p.extintor_id) if p.extintor_id else None
    if ext and ext.situacao == "ATENCAO":
        ext.situacao = "NO_PRAZO"
    _log("Extintor", f"Pendência de etiqueta baixada: {p.extintor_cod} ({p.local})")
    db.session.commit()
    flash("Pendência de etiqueta resolvida.", "success")
    return redirect(url_for("almox.pendencias_etiqueta"))


def _extintores_filtrados():
    """Lista de extintores ativos aplicando os mesmos filtros da tela (predio/local/tipo/situacao/ids)."""
    predios_sel = _args_list("predio")
    locais_sel = _args_list("local")
    tipos_sel = _args_list("tipo")
    situacoes = _situacoes_sel()
    ids = request.args.get("ids") or ""
    consulta = Extintor.query.filter_by(ativo=True)
    if ids.strip():
        lista_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        consulta = consulta.filter(Extintor.id.in_(lista_ids))
    out = []
    for e in consulta.order_by(Extintor.predio, Extintor.local, Extintor.codigo).all():
        if predios_sel and e.predio not in predios_sel:
            continue
        if tipos_sel and e.tipo not in tipos_sel:
            continue
        if locais_sel and e.local not in locais_sel:
            continue
        if not _match_situacoes(situacoes, _situacao_extintor(e)[0]):
            continue
        out.append(e)
    return out


def _filtro_atual():
    """Dict com os filtros ativos, p/ repassar em links (mantem a selecao)."""
    d = {}
    for k in ("predio", "local", "tipo", "situacao"):
        v = _args_list(k)
        if v:
            d[k] = v
    if request.args.get("ids"):
        d["ids"] = request.args.get("ids")
    return d


@almox_bp.route("/extintores/qr")
@_guard("pode_extintores")
def extintores_qr():
    itens = _extintores_filtrados()
    formato = request.args.get("formato", "termica")
    f = _filtro_atual()
    return render_template("almox/extintores_qr.html", itens=itens, qr_svg=_qr_svg,
                           formato=formato,
                           f_predio=f.get("predio", ""), f_local=f.get("local", ""),
                           f_tipo=f.get("tipo", ""), f_situacao=f.get("situacao", ""),
                           f_ids=f.get("ids", ""))


@almox_bp.route("/extintores/etiquetas.pdf")
@_guard("pode_extintores")
def extintores_etiquetas_pdf():
    """PDF no tamanho EXATO do rolo de etiquetas (2 colunas de 45x20mm). Cada pagina = uma linha
    (as duas etiquetas lado a lado). Impressao 100%, sem escala -> encaixa em cada etiqueta.
    Medidas ajustaveis por querystring (mm): lw, lh, ml, gap, mr, rowgap, qr, dx, dy."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from flask import Response
    import io, qrcode

    def _f(k, d):
        try:
            return float(request.args.get(k, d))
        except (TypeError, ValueError):
            return d

    itens = _extintores_filtrados()

    # Geometria (mm) — medidas reais do rolo (ajustaveis via querystring)
    LW = _f("lw", 45.0)      # largura etiqueta
    LH = _f("lh", 20.0)      # altura etiqueta
    ML = _f("ml", 2.0)       # margem esquerda
    GAP = _f("gap", 2.0)     # vao entre as duas colunas
    MR = _f("mr", 1.0)       # margem direita
    ROWGAP = _f("rowgap", 3.0)   # vao entre linhas (gap do sensor)
    QR = _f("qr", 16.0)      # tamanho do QR
    DX = _f("dx", 0.0) * mm  # ajuste fino do conteudo X
    DY = _f("dy", 0.0) * mm  # ajuste fino do conteudo Y
    modo = request.args.get("modo", "gap")   # 'gap' (pagina = etiqueta) ou 'continuo' (inclui rowgap)

    LINER = ML + LW + GAP + LW + MR          # 95
    PW = LINER * mm
    PH = (LH + (ROWGAP if modo == "continuo" else 0)) * mm
    base = (ROWGAP * mm) if modo == "continuo" else 0   # no continuo, gap fica embaixo

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PW, PH))
    host = request.host_url
    colx = [ML * mm, (ML + LW + GAP) * mm]

    _cache = {}
    def _matrix(url):
        if url not in _cache:
            q = qrcode.QRCode(border=0, error_correction=qrcode.constants.ERROR_CORRECT_M)
            q.add_data(url); q.make(fit=True)
            _cache[url] = q.get_matrix()
        return _cache[url]

    def wrap(s, n):
        out, cur = [], ""
        for w in (s or "").split():
            if len(cur) + len(w) + (1 if cur else 0) <= n:
                cur = (cur + " " + w).strip()
            else:
                out.append(cur); cur = w
        if cur:
            out.append(cur)
        return out or [""]

    def draw_label(x0, e):
        url = host + "almoxarifado/e/" + (e.qr_uid or str(e.id))
        qrsize = QR * mm
        qx = x0 + 1.5 * mm + DX
        qy = base + (LH * mm - qrsize) / 2 + DY
        m = _matrix(url); n = len(m); cell = qrsize / n
        c.setFillGray(0)
        for r in range(n):           # desenha por linhas, juntando sequencias (run-length) -> leve
            row = m[r]; col = 0; yy = qy + (n - 1 - r) * cell
            while col < n:
                if row[col]:
                    ini = col
                    while col < n and row[col]:
                        col += 1
                    c.rect(qx + ini * cell, yy, (col - ini) * cell, cell, stroke=0, fill=1)
                else:
                    col += 1
        tx = qx + qrsize + 1.6 * mm
        top = base + LH * mm - 4.2 * mm + DY
        c.setFont("Helvetica-Bold", 9)
        c.drawString(tx, top, (e.codigo or "-"))
        c.setFont("Helvetica", 6)
        loc = " · ".join([p for p in [e.predio, e.local] if p])
        yy = top - 3.1 * mm
        for ln in wrap(loc, 24)[:2]:
            c.drawString(tx, yy, ln); yy -= 2.5 * mm
        c.setFont("Helvetica", 4.5)
        c.drawString(tx, base + 3.0 * mm + DY, "SERENA · CLUSTER DELTA")

    i = 0
    if not itens:
        c.showPage()
    while i < len(itens):
        for col in range(2):
            if i < len(itens):
                draw_label(colx[col], itens[i]); i += 1
        c.showPage()
        c.setPageSize((PW, PH))
    c.save()
    buf.seek(0)
    return Response(buf.read(), mimetype="application/pdf",
                    headers={"Content-Disposition": "inline; filename=etiquetas_extintores.pdf"})


@almox_bp.route("/extintores/pdf")
@_guard("pode_extintores")
def extintores_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io
    from flask import Response
    predios_sel = _args_list("predio")
    locais_sel = _args_list("local")
    tipos_sel = _args_list("tipo")
    situacoes = _situacoes_sel()
    ids = request.args.get("ids") or ""
    consulta = Extintor.query.filter_by(ativo=True)
    if ids.strip():
        consulta = consulta.filter(Extintor.id.in_([int(x) for x in ids.split(",") if x.strip().isdigit()]))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12 * mm, bottomMargin=12 * mm,
                            leftMargin=10 * mm, rightMargin=10 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Extintores — " + (", ".join(predios_sel) if predios_sel else "Todos os prédios"), styles["Title"]), Spacer(1, 6)]
    data = [["Código", "Prédio", "Local", "Tipo/Carga", "Classe", "Validade", "TH", "Situação"]]
    resumo = {}
    for e in consulta.order_by(Extintor.predio, Extintor.local, Extintor.codigo).all():
        if predios_sel and e.predio not in predios_sel:
            continue
        if tipos_sel and e.tipo not in tipos_sel:
            continue
        if locais_sel and e.local not in locais_sel:
            continue
        k, lbl, _ = _situacao_extintor(e)
        if not _match_situacoes(situacoes, k):
            continue
        chave = (e.tipo or "—", e.classe or "—")
        resumo[chave] = resumo.get(chave, 0) + 1
        data.append([e.codigo or "", e.predio or "", e.local or "", e.tipo or "", e.classe or "",
                     _competencia(e.validade), _th_label(e.teste_hidrostatico), lbl])
    # Somatorio por tipo/carga + classe (antes do relatorio completo) — ajuda o prestador a se organizar
    total_ext = sum(resumo.values())
    res_data = [["Tipo/Carga", "Classe", "Qtd"]]
    for (tp, cl), q in sorted(resumo.items()):
        res_data.append([tp, cl, str(q)])
    res_data.append(["", "TOTAL", str(total_ext)])
    res_tbl = Table(res_data, repeatRows=1, colWidths=[70 * mm, 34 * mm, 22 * mm])
    res_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4B4B4B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EDE9E5")),
    ]))
    elems.append(Paragraph("Resumo por tipo/carga e classe", styles["Heading3"]))
    elems.append(res_tbl)
    elems.append(Spacer(1, 8))
    elems.append(Paragraph("Relatório completo", styles["Heading3"]))
    elems.append(Spacer(1, 4))
    if len(data) == 1:
        data.append(["—"] * 8)
    t = Table(data, repeatRows=1,
              colWidths=[24 * mm, 22 * mm, 60 * mm, 34 * mm, 16 * mm, 24 * mm, 24 * mm, 40 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5246")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=extintores.pdf"})



# ---------- COLABORADORES ----------
PAPEIS_COLAB = [("COLABORADOR DIVERSO", "Colaborador diverso"),
                ("SOLICITANTE", "Solicitante"),
                ("ALMOXARIFADO", "Almoxarifado"),
                ("VISUALIZADOR", "Visualizador"),
                ("ADMIN", "Admin")]


@almox_bp.route("/colaboradores")
@_guard("pode_colaboradores")
def colaboradores():
    from .models import Fornecedor
    itens = Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome).all()
    papeis = PapelColaborador.query.filter_by(ativo=True).order_by(PapelColaborador.nome).all()
    empresas = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).all()
    return render_template("almox/colaboradores.html", itens=itens, papeis=papeis, empresas=empresas,
                           papeis_colab=PAPEIS_COLAB, pode_papel=current_user.is_admin,
                           is_master=current_user.is_master)


@almox_bp.route("/colaboradores/novo", methods=["POST"])
@_guard("pode_colaboradores")
def colaborador_novo():
    import secrets
    nome = (request.form.get("nome") or "").strip()
    cpf_bruto = (request.form.get("cpf") or "").strip()
    cpf = "".join(ch for ch in cpf_bruto if ch.isdigit())
    email = (request.form.get("email") or "").strip().lower()
    if not nome:
        flash("Informe o nome completo do colaborador.", "danger")
        return redirect(url_for("almox.colaboradores"))
    if not cpf:
        flash("CPF é obrigatório (só números).", "danger")
        return redirect(url_for("almox.colaboradores"))
    for c0 in Colaborador.query.filter(Colaborador.ativo.is_(True)).all():
        if "".join(ch for ch in (c0.cpf or "") if ch.isdigit()) == cpf:
            flash("Já existe um colaborador ativo com esse CPF.", "warning")
            return redirect(url_for("almox.colaboradores"))
    # Perfil de acesso: só Admin escolhe; senão entra sem perfil administrativo.
    papel = "COLABORADOR DIVERSO"
    if current_user.is_admin:
        escolhido = (request.form.get("papel") or "").strip().upper()
        nomes = {p.nome.upper() for p in PapelColaborador.query.all()}
        if escolhido and escolhido in nomes:
            papel = escolhido
    uid = "COL-" + secrets.token_hex(4).upper()
    while Colaborador.query.filter_by(qr_uid=uid).first():
        uid = "COL-" + secrets.token_hex(4).upper()
    c = Colaborador(nome=nome.upper(), cpf=cpf, email=(email or None),
                    empresa=(request.form.get("empresa") or "").strip().upper(),
                    cargo=(request.form.get("cargo") or "").strip().upper(),
                    papel=papel, qr_uid=uid)
    db.session.add(c)
    _log("Colaborador", f"Colaborador cadastrado: {c.nome} (CPF {cpf}, papel {papel})")
    db.session.commit()
    flash("Colaborador cadastrado. Ele define a senha no 1º acesso (CPF).", "success")
    return redirect(url_for("almox.colaboradores"))


@almox_bp.route("/colaboradores/<int:cid>", methods=["GET"])
@_guard("pode_colaboradores")
def colaborador_perfil(cid):
    from .models import Fornecedor, HistoricoColaborador
    c = db.session.get(Colaborador, cid) or abort(404)
    empresas = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).all()
    hist = (HistoricoColaborador.query.filter_by(colaborador_id=c.id)
            .order_by(HistoricoColaborador.criado_em.desc()).all())
    papeis = PapelColaborador.query.filter_by(ativo=True).order_by(PapelColaborador.nome).all()
    return render_template("almox/colaborador_perfil.html", c=c, empresas=empresas,
                           papeis=papeis, papeis_colab=PAPEIS_COLAB, hist=hist,
                           pode_papel=current_user.is_admin, is_master=current_user.is_master)


@almox_bp.route("/colaboradores/<int:cid>/editar", methods=["POST"])
@_guard("pode_colaboradores")
def colaborador_editar(cid):
    from .models import HistoricoColaborador
    c = db.session.get(Colaborador, cid) or abort(404)
    autor = current_user.nome
    mudancas = []

    def registrar(campo, de, para):
        db.session.add(HistoricoColaborador(colaborador_id=c.id, colaborador_nome=c.nome,
                       campo=campo, de=de or "—", para=para or "—", alterado_por_nome=autor))
        mudancas.append(campo)

    # Cargo e empresa: Admin/Master e Almoxarifado podem
    novo_cargo = (request.form.get("cargo") or "").strip().upper()
    if novo_cargo != (c.cargo or ""):
        registrar("cargo", c.cargo, novo_cargo); c.cargo = novo_cargo
    nova_emp = (request.form.get("empresa") or "").strip().upper()
    if nova_emp != (c.empresa or ""):
        registrar("empresa", c.empresa, nova_emp); c.empresa = nova_emp

    # Perfil de acesso: só Admin
    if current_user.is_admin:
        novo_papel = (request.form.get("papel") or c.papel or "").strip().upper()
        nomes = {p.nome.upper() for p in PapelColaborador.query.all()}
        if novo_papel and novo_papel in nomes and novo_papel != (c.papel or "").upper():
            registrar("papel", c.papel, novo_papel); c.papel = novo_papel
    elif request.form.get("papel") and request.form.get("papel").strip().upper() != (c.papel or "").upper():
        flash("Apenas Admin altera o perfil de acesso. As demais alterações foram salvas.", "warning")

    if mudancas:
        _log("Colaborador", f"{c.nome}: alterado ({', '.join(mudancas)}) por {autor}")
    db.session.commit()
    flash("Alterações salvas." if mudancas else "Nada a alterar.", "success" if mudancas else "info")
    return redirect(url_for("almox.colaborador_perfil", cid=c.id))


@almox_bp.route("/colaboradores/<int:cid>/desativar", methods=["POST"])
@_guard("pode_colaboradores")
def colaborador_desativar(cid):
    c = db.session.get(Colaborador, cid) or abort(404)
    c.ativo = False
    _log("Colaborador", f"Colaborador desativado: {c.nome}")
    db.session.commit()
    flash("Colaborador desativado.", "success")
    return redirect(url_for("almox.colaboradores"))


@almox_bp.route("/colaboradores/<int:cid>/reset-senha", methods=["POST"])
@_guard("pode_colaboradores")
def colaborador_reset_senha(cid):
    c = db.session.get(Colaborador, cid) or abort(404)
    c.senha_hash = None
    _log("Colaborador", f"Senha resetada: {c.nome}")
    db.session.commit()
    flash(f"Senha de {c.nome} resetada. No próximo uso ele define uma nova.", "success")
    return redirect(url_for("almox.colaboradores"))


@almox_bp.route("/colaboradores/qr")
@_guard("pode_colaboradores")
def colaboradores_qr():
    ids = request.args.get("ids") or ""
    consulta = Colaborador.query.filter_by(ativo=True)
    if ids.strip():
        lista_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        consulta = consulta.filter(Colaborador.id.in_(lista_ids))
    itens = consulta.order_by(Colaborador.nome).all()
    formato = request.args.get("formato", "foto")   # 'foto' | 'termica'
    return render_template("almox/colaboradores_qr.html", itens=itens, qr_svg=_qr_svg, formato=formato)


def _admin_only(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return wrapper


# ---------- PAPÉIS DE COLABORADOR (somente ADMIN) ----------
@almox_bp.route("/papeis")
@_admin_only
def papeis():
    itens = PapelColaborador.query.order_by(PapelColaborador.nome).all()
    return render_template("almox/papeis.html", itens=itens,
                           tarefas_perfil=TAREFAS_PERFIL, grupos=TAREFAS_GRUPOS)


@almox_bp.route("/papeis/novo", methods=["POST"])
@_admin_only
def papel_novo():
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        flash("Informe o nome do papel.", "danger")
        return redirect(url_for("almox.papeis"))
    if PapelColaborador.query.filter(db.func.upper(PapelColaborador.nome) == nome.upper()).first():
        flash("Já existe um papel com esse nome.", "warning")
        return redirect(url_for("almox.papeis"))
    validas = {k for k, _r, _g, fut in TAREFAS_PERFIL if not fut} | {k for k, _ in TAREFAS_COLABORADOR}
    escolhidas = [t for t in request.form.getlist("tarefas") if t in validas]
    db.session.add(PapelColaborador(nome=nome.upper(), tarefas=",".join(escolhidas)))
    _log("Papel", f"Papel cadastrado: {nome.upper()}")
    db.session.commit()
    flash("Papel cadastrado.", "success")
    return redirect(url_for("almox.papeis"))


@almox_bp.route("/papeis/<int:pid>/editar", methods=["POST"])
@_admin_only
def papel_editar(pid):
    p = db.session.get(PapelColaborador, pid) or abort(404)
    validas = {k for k, _r, _g, fut in TAREFAS_PERFIL if not fut} | {k for k, _ in TAREFAS_COLABORADOR}
    p.tarefas = ",".join(t for t in request.form.getlist("tarefas") if t in validas)
    _log("Papel", f"Papel editado: {p.nome}")
    db.session.commit()
    flash("Papel atualizado.", "success")
    return redirect(url_for("almox.papeis"))


@almox_bp.route("/papeis/<int:pid>/toggle", methods=["POST"])
@_admin_only
def papel_toggle(pid):
    p = db.session.get(PapelColaborador, pid) or abort(404)
    p.ativo = not p.ativo
    db.session.commit()
    return redirect(url_for("almox.papeis"))


# ---------- RELATÓRIO DE CHAVES ----------
def _filtra_movimentacoes():
    """Aplica os filtros da querystring e devolve (lista, contexto_filtros)."""
    from datetime import datetime as _dt
    qy = MovimentacaoChave.query
    di = request.args.get("data_ini") or ""
    dfim = request.args.get("data_fim") or ""
    chave_id = request.args.get("chave_id") or ""
    colaborador = (request.args.get("colaborador") or "").strip()
    quadro = (request.args.get("quadro") or "").strip()
    acao = request.args.get("acao") or ""
    if di:
        try: qy = qy.filter(MovimentacaoChave.criado_em >= _dt.strptime(di, "%Y-%m-%d"))
        except ValueError: pass
    if dfim:
        try:
            fim = _dt.strptime(dfim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            qy = qy.filter(MovimentacaoChave.criado_em <= fim)
        except ValueError: pass
    if chave_id.isdigit():
        qy = qy.filter(MovimentacaoChave.chave_id == int(chave_id))
    if acao in ("retirada", "devolucao"):
        qy = qy.filter(MovimentacaoChave.acao == acao)
    movs = qy.order_by(MovimentacaoChave.criado_em.desc()).all()
    if colaborador:
        movs = [m for m in movs if contem_busca(m.colaborador_nome, colaborador)]
    if quadro:
        movs = [m for m in movs if contem_busca(m.quadro_nome, quadro)]
    ctx = dict(data_ini=di, data_fim=dfim, chave_id=chave_id, colaborador=colaborador,
               quadro=quadro, acao=acao)
    return movs, ctx


@almox_bp.route("/chaves/relatorio")
@_guard("pode_chaves")
def relatorio_chaves():
    movs, ctx = _filtra_movimentacoes()
    # Resumos
    retiradas = [m for m in movs if m.acao == "retirada"]
    total_retiradas = len(retiradas)
    colabs_distintos = len({(m.colaborador_nome or "").upper() for m in retiradas})
    # ranking chaves e colaboradores (por retiradas)
    from collections import Counter
    rk_chaves = Counter((m.chave_desc or "—") for m in retiradas).most_common(10)
    rk_colabs = Counter((m.colaborador_nome or "—") for m in retiradas).most_common(10)
    # situação atual: chaves em uso
    em_uso = Chave.query.filter_by(ativo=True, status="Em uso").order_by(Chave.descricao).all()
    chaves = Chave.query.filter_by(ativo=True).order_by(Chave.descricao).all()
    return render_template("almox/relatorio_chaves.html", movs=movs, ctx=ctx,
                           total_retiradas=total_retiradas, colabs_distintos=colabs_distintos,
                           rk_chaves=rk_chaves, rk_colabs=rk_colabs, em_uso=em_uso, chaves=chaves)


@almox_bp.route("/chaves/relatorio/csv")
@_guard("pode_chaves")
def relatorio_chaves_csv():
    import csv, io
    movs, _ = _filtra_movimentacoes()
    buf = io.StringIO()
    buf.write("\ufeff")   # BOM p/ Excel abrir acentos certo
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Data/Hora", "Ação", "Chave", "Quadro", "Colaborador", "Estava com (devolução)", "Operador"])
    for m in movs:
        w.writerow([m.criado_em.strftime("%d/%m/%Y %H:%M") if m.criado_em else "",
                    m.acao, m.chave_desc or "", m.quadro_nome or "",
                    m.colaborador_nome or "",
                    (m.retirado_por or "") if m.acao == "devolucao" else "",
                    (m.operador.nome if m.operador else "")])
    from flask import Response
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=relatorio_chaves.csv"})


@almox_bp.route("/chaves/relatorio/pdf")
@_guard("pode_chaves")
def relatorio_chaves_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io
    from flask import Response
    movs, ctx = _filtra_movimentacoes()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Relatório de Movimentações de Chaves", styles["Title"])]
    periodo = f"Período: {ctx['data_ini'] or 'início'} a {ctx['data_fim'] or 'hoje'}"
    elems.append(Paragraph(periodo, styles["Normal"]))
    elems.append(Spacer(1, 6))
    data = [["Data/Hora", "Ação", "Chave", "Quadro", "Colaborador", "Estava com", "Operador"]]
    for m in movs:
        data.append([m.criado_em.strftime("%d/%m/%y %H:%M") if m.criado_em else "",
                     m.acao, m.chave_desc or "", m.quadro_nome or "",
                     m.colaborador_nome or "",
                     (m.retirado_por or "") if m.acao == "devolucao" else "",
                     (m.operador.nome if m.operador else "")])
    if len(data) == 1:
        data.append(["—", "—", "—", "—", "—", "—", "—"])
    t = Table(data, repeatRows=1, colWidths=[22*mm, 17*mm, 33*mm, 33*mm, 30*mm, 30*mm, 22*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5246")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=relatorio_chaves.pdf"})


# ==================== COLETOR (online — fase 1) ====================
# Aparelho dedicado do almoxarifado (logado como ALMOXARIFADO). Lê QR pela câmera.
# Identifica o colaborador por QR + senha e movimenta CHAVES. Material entra depois.
from flask import jsonify


@almox_bp.route("/coletor")
@_guard("pode_coletor")
def coletor():
    abas = {
        "chaves": _efetivo("col_chaves"),
        "material": _efetivo("col_material"),
        "mover": _efetivo("col_movimentacao"),
        "inventario": _efetivo("col_inventario"),
    }
    return render_template("almox/coletor.html", abas=abas)


def _aba_coletor_ok(aba):
    """True se o usuário (ou perfil simulado) pode a aba do coletor: chaves|material|mover|inventario."""
    mapa = {"chaves": "col_chaves", "material": "col_material",
            "mover": "col_movimentacao", "inventario": "col_inventario"}
    return _efetivo(mapa.get(aba, ""))


def _aba_material_qualquer():
    """Qualquer aba que lida com material/localizador (material, mover ou inventário)."""
    return _aba_coletor_ok("material") or _aba_coletor_ok("mover") or _aba_coletor_ok("inventario")


def _nega_aba(aba, msg=None):
    textos = {
        "chaves": "Seu perfil não pode usar o coletor para chaves.",
        "material": "Seu perfil não pode usar o coletor para material.",
        "mover": "Seu perfil não pode movimentar entre localizadores.",
        "inventario": "Seu perfil não pode fazer inventário pelo coletor.",
    }
    return jsonify(ok=False, erro=msg or textos.get(aba, "Ação não permitida para o seu perfil.")), 403


def _uid_limpo(qr_uid):
    """Aceita tanto 'CHAVE:CH-XXXX' / 'COLAB:COL-XXXX' quanto só o uid."""
    s = (qr_uid or "").strip()
    if ":" in s:
        s = s.split(":", 1)[1]
    return s


@almox_bp.route("/coletor/api/colaborador/<path:qr_uid>")
@modulo_required
def coletor_api_colaborador(qr_uid):
    c = _resolver_colaborador(_uid_limpo(qr_uid)) or _resolver_colaborador(qr_uid)
    if not c:
        return jsonify(ok=False, erro="Colaborador não encontrado."), 404
    return jsonify(ok=True, id=c.id, nome=c.nome, tem_senha=c.tem_senha)


@almox_bp.route("/coletor/api/chave/<path:qr_uid>")
@modulo_required
def coletor_api_chave(qr_uid):
    if not _aba_coletor_ok("chaves"):
        return _nega_aba("chaves")
    uid = _uid_limpo(qr_uid)
    c = Chave.query.filter_by(qr_uid=uid, ativo=True).first()
    if not c:
        return jsonify(ok=False, erro="Chave não encontrada."), 404
    return jsonify(ok=True, id=c.id, descricao=c.descricao, quadro=c.quadro_nome,
                   status=c.status, com_quem=c.com_quem)


@almox_bp.route("/coletor/api/confirmar", methods=["POST"])
@modulo_required
def coletor_api_confirmar():
    if not _aba_coletor_ok("chaves"): return _nega_aba("chaves")
    """Aplica as ações de chave confirmadas pela senha do colaborador.
    payload: {colaborador_id, senha, acoes:[{chave_id, acao: 'retirar'|'devolver'}]}"""
    data = request.get_json(silent=True) or {}
    colab = db.session.get(Colaborador, data.get("colaborador_id") or 0)
    if not colab or not colab.ativo:
        return jsonify(ok=False, erro="Colaborador inválido."), 400
    senha = data.get("senha") or ""
    # Primeiro uso define a senha; depois valida.
    if not colab.tem_senha:
        if len(senha) < 4:
            return jsonify(ok=False, erro="Primeiro uso: defina uma senha de ao menos 4 dígitos."), 400
        colab.set_senha(senha)
    elif not colab.check_senha(senha):
        return jsonify(ok=False, erro="Senha do colaborador inválida."), 403

    acoes = data.get("acoes") or []
    if not acoes:
        return jsonify(ok=False, erro="Nenhuma chave para confirmar."), 400
    feitas = []
    for a in acoes:
        c = db.session.get(Chave, a.get("chave_id") or 0)
        if not c or not c.ativo:
            continue
        acao = a.get("acao")
        if acao == "retirar" and c.status == "Disponível":
            c.status = "Em uso"
            c.com_quem = colab.nome
            db.session.add(MovimentacaoChave(chave_id=c.id, chave_desc=c.descricao,
                           quadro_nome=c.quadro_nome, colaborador_id=colab.id,
                           colaborador_nome=colab.nome, acao="retirada",
                           operador_id=_op_id()))
            feitas.append(f"Retirou {c.descricao}")
        elif acao == "devolver" and c.status == "Em uso":
            retirou = c.com_quem or "—"
            c.status = "Disponível"
            c.com_quem = None
            db.session.add(MovimentacaoChave(chave_id=c.id, chave_desc=c.descricao,
                           quadro_nome=c.quadro_nome, colaborador_id=colab.id,
                           colaborador_nome=colab.nome, retirado_por=retirou,
                           acao="devolucao", operador_id=_op_id()))
            feitas.append(f"Devolveu {c.descricao}")
    _log("Coletor", f"{colab.nome}: " + "; ".join(feitas) if feitas else f"{colab.nome}: sem ações válidas")
    db.session.commit()
    return jsonify(ok=True, resumo=feitas, colaborador=colab.nome)


# ==================== MATERIAL (estoque com quantidade) ====================
def _num(v, default=0.0):
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return default


def _filtra_saldo():
    """Filtros do relatório de saldo (compartilhado tela + CSV): categoria, localizador, abaixo do mínimo."""
    cat = (request.args.get("cat") or "").strip()
    loc = request.args.get("loc") or ""
    baixo = request.args.get("baixo") == "1"
    q = ProdutoAlmox.query.filter_by(ativo=True)
    itens = q.order_by(ProdutoAlmox.nome).all()
    if cat:
        itens = [p for p in itens if contem_busca(p.categoria, cat)]
    if loc.isdigit():
        ids = {el.produto_id for el in EstoqueLocalizador.query.filter(
            EstoqueLocalizador.localizador_id == int(loc),
            EstoqueLocalizador.quantidade > 0).all()}
        itens = [p for p in itens if p.id in ids]
    if baixo:
        itens = [p for p in itens if p.abaixo_minimo]
    ctx = dict(cat=cat, loc=loc, baixo=("1" if baixo else ""))
    return itens, ctx


@almox_bp.route("/materiais")
@_guard("pode_material")
def materiais():
    itens, ctx = _filtra_saldo()
    n_baixo = sum(1 for p in itens if p.abaixo_minimo)
    locais = LocalAlmox.query.filter_by(ativo=True).order_by(LocalAlmox.nome).all()
    fabricantes = Fabricante.query.filter_by(ativo=True).order_by(Fabricante.nome).all()
    from .models import Localizador
    localizadores = Localizador.query.filter_by(ativo=True).all()
    localizadores.sort(key=lambda l: l.codigo)
    categorias = sorted({(p.categoria or "").strip() for p in ProdutoAlmox.query.filter_by(ativo=True).all() if (p.categoria or "").strip()})
    return render_template("almox/materiais.html", itens=itens, n_baixo=n_baixo, locais=locais,
                           fabricantes=fabricantes, ctx=ctx, localizadores=localizadores, categorias=categorias)


@almox_bp.route("/materiais/novo", methods=["POST"])
@_guard("pode_material")
def material_novo():
    import secrets
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        flash("Informe o nome do material.", "danger")
        return redirect(url_for("almox.materiais"))
    uid = "MAT-" + secrets.token_hex(4).upper()
    while ProdutoAlmox.query.filter_by(qr_uid=uid).first():
        uid = "MAT-" + secrets.token_hex(4).upper()
    local_id = request.form.get("local_id") or None
    if not local_id:
        temp = LocalAlmox.query.filter_by(temporaria=True).first()
        local_id = temp.id if temp else None
    p = ProdutoAlmox(codigo=(request.form.get("codigo") or "").strip().upper(),
                     codigo_barras=(request.form.get("codigo_barras") or "").strip() or None,
                     nome=nome.upper(), unidade=(request.form.get("unidade") or "UN").strip().upper(),
                     categoria=(request.form.get("categoria") or "").strip().upper(),
                     saldo=_num(request.form.get("saldo_inicial"), 0),
                     saldo_minimo=_num(request.form.get("saldo_minimo"), 0),
                     fabricante_id=int(request.form["fabricante_id"]) if request.form.get("fabricante_id") else None,
                     opc_tag=bool(request.form.get("opc_tag")),
                     opc_ca=bool(request.form.get("opc_ca")),
                     opc_validade=bool(request.form.get("opc_validade")),
                     opc_validade_calib=bool(request.form.get("opc_validade_calib")),
                     opc_lote=bool(request.form.get("opc_lote")),
                     local_id=int(local_id) if local_id else None, qr_uid=uid)
    db.session.add(p)
    db.session.flush()
    if p.saldo:
        db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="entrada",
                       quantidade=p.saldo, saldo_apos=p.saldo, operador_id=_op_id(),
                       obs="Saldo inicial de cadastro"))
    _log("Material", f"Material cadastrado: {p.nome} (saldo {p.saldo} {p.unidade})")
    db.session.commit()
    flash("Material cadastrado.", "success")
    return redirect(url_for("almox.materiais"))


# ----- Locais de estocagem -----
@almox_bp.route("/materiais/locais")
@_guard("pode_material")
def locais():
    itens = LocalAlmox.query.order_by(LocalAlmox.nome).all()
    return render_template("almox/locais.html", itens=itens)


@almox_bp.route("/materiais/locais/novo", methods=["POST"])
@_guard("pode_material")
def local_novo():
    nome = (request.form.get("nome") or "").strip()
    if not nome:
        flash("Informe o nome do local.", "danger")
        return redirect(url_for("almox.locais"))
    if LocalAlmox.query.filter(db.func.upper(LocalAlmox.nome) == nome.upper()).first():
        flash("Já existe um local com esse nome.", "warning")
        return redirect(url_for("almox.locais"))
    db.session.add(LocalAlmox(nome=nome.upper()))
    _log("Material", f"Local cadastrado: {nome.upper()}")
    db.session.commit()
    flash("Local cadastrado.", "success")
    return redirect(url_for("almox.locais"))


@almox_bp.route("/materiais/locais/<int:lid>/toggle", methods=["POST"])
@_guard("pode_material")
def local_toggle(lid):
    l = db.session.get(LocalAlmox, lid) or abort(404)
    l.ativo = not l.ativo
    db.session.commit()
    return redirect(url_for("almox.locais"))


@almox_bp.route("/materiais/<int:pid>/mover", methods=["POST"])
@_guard("pode_material")
def material_mover(pid):
    p = db.session.get(ProdutoAlmox, pid) or abort(404)
    destino_id = request.form.get("local_id")
    destino = db.session.get(LocalAlmox, int(destino_id)) if destino_id and destino_id.isdigit() else None
    if not destino:
        flash("Selecione o local de destino.", "danger")
        return redirect(url_for("almox.materiais"))
    de = p.local_nome
    p.local_id = destino.id
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="movimentacao",
                   quantidade=0, saldo_apos=p.saldo, local_de=de, local_para=destino.nome,
                   operador_id=_op_id(), obs=(request.form.get("obs") or "").strip()))
    _log("Material", f"{p.nome}: movido de {de} para {destino.nome}")
    db.session.commit()
    flash(f"{p.nome} movido para {destino.nome}.", "success")
    return redirect(url_for("almox.materiais"))


@almox_bp.route("/materiais/<int:pid>/entrada", methods=["POST"])
@_guard("pode_material")
def material_entrada(pid):
    p = db.session.get(ProdutoAlmox, pid) or abort(404)
    qtd = _num(request.form.get("quantidade"))
    if qtd <= 0:
        flash("Quantidade inválida.", "danger")
        return redirect(url_for("almox.materiais"))
    p.saldo = (p.saldo or 0) + qtd
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="entrada",
                   quantidade=qtd, saldo_apos=p.saldo, operador_id=_op_id(),
                   obs=(request.form.get("obs") or "").strip()))
    _log("Material", f"Entrada {qtd} {p.unidade} de {p.nome} (saldo {p.saldo})")
    db.session.commit()
    flash(f"Entrada registrada. Saldo de {p.nome}: {p.saldo:g} {p.unidade}.", "success")
    return redirect(url_for("almox.materiais"))


@almox_bp.route("/materiais/<int:pid>/saida", methods=["POST"])
@_guard("pode_material")
def material_saida(pid):
    p = db.session.get(ProdutoAlmox, pid) or abort(404)
    qtd = _num(request.form.get("quantidade"))
    if qtd <= 0:
        flash("Quantidade inválida.", "danger")
        return redirect(url_for("almox.materiais"))
    # Prevenção de estoque negativo (item roadmap §9)
    if qtd > (p.saldo or 0):
        flash(f"Saída bloqueada: saldo de {p.nome} é {p.saldo:g} {p.unidade}, "
              f"menor que {qtd:g}. Faça entrada/ajuste antes.", "danger")
        return redirect(url_for("almox.materiais"))
    nome = (request.form.get("colaborador_nome") or "").strip().upper()
    colab = Colaborador.query.filter(db.func.upper(Colaborador.nome) == nome).first() if nome else None
    p.saldo = (p.saldo or 0) - qtd
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="saida",
                   quantidade=qtd, saldo_apos=p.saldo, colaborador_id=(colab.id if colab else None),
                   colaborador_nome=(nome or None), operador_id=_op_id(),
                   obs=(request.form.get("obs") or "").strip()))
    _log("Material", f"Saída {qtd} {p.unidade} de {p.nome} p/ {nome or '—'} (saldo {p.saldo})")
    db.session.commit()
    flash(f"Saída registrada. Saldo de {p.nome}: {p.saldo:g} {p.unidade}.", "success")
    return redirect(url_for("almox.materiais"))


@almox_bp.route("/materiais/<int:pid>/ajuste", methods=["POST"])
@_guard("pode_material")
def material_ajuste(pid):
    p = db.session.get(ProdutoAlmox, pid) or abort(404)
    novo = _num(request.form.get("saldo"))
    antigo = p.saldo or 0
    p.saldo = novo
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="ajuste",
                   quantidade=novo - antigo, saldo_apos=novo, operador_id=_op_id(),
                   obs=(request.form.get("obs") or f"Ajuste de {antigo:g} para {novo:g}").strip()))
    _log("Material", f"Ajuste de saldo {p.nome}: {antigo:g} → {novo:g}")
    db.session.commit()
    flash(f"Saldo de {p.nome} ajustado para {novo:g} {p.unidade}.", "success")
    return redirect(url_for("almox.materiais"))


@almox_bp.route("/materiais/<int:pid>/desativar", methods=["POST"])
@_guard("pode_material")
def material_desativar(pid):
    p = db.session.get(ProdutoAlmox, pid) or abort(404)
    p.ativo = False
    _log("Material", f"Material desativado: {p.nome}")
    db.session.commit()
    flash("Material desativado.", "success")
    return redirect(url_for("almox.materiais"))


def _filtra_mov_material():
    from datetime import datetime as _dt
    qy = MovimentacaoMaterial.query
    pid = request.args.get("produto_id") or ""
    tipo = request.args.get("tipo") or ""
    di = request.args.get("data_ini") or ""
    dfim = request.args.get("data_fim") or ""
    if pid.isdigit():
        qy = qy.filter(MovimentacaoMaterial.produto_id == int(pid))
    if tipo in ("entrada", "saida", "ajuste"):
        qy = qy.filter(MovimentacaoMaterial.tipo == tipo)
    if di:
        try: qy = qy.filter(MovimentacaoMaterial.criado_em >= _dt.strptime(di, "%Y-%m-%d"))
        except ValueError: pass
    if dfim:
        try:
            fim = _dt.strptime(dfim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            qy = qy.filter(MovimentacaoMaterial.criado_em <= fim)
        except ValueError: pass
    movs = qy.order_by(MovimentacaoMaterial.criado_em.desc()).limit(2000).all()
    ctx = dict(produto_id=pid, tipo=tipo, data_ini=di, data_fim=dfim)
    return movs, ctx


@almox_bp.route("/materiais/movimentacoes")
@_guard("pode_material")
def materiais_mov():
    movs, ctx = _filtra_mov_material()
    produtos = ProdutoAlmox.query.filter_by(ativo=True).order_by(ProdutoAlmox.nome).all()
    return render_template("almox/materiais_mov.html", movs=movs, produtos=produtos, ctx=ctx)


@almox_bp.route("/materiais/movimentacoes/csv")
@_guard("pode_material")
def materiais_mov_csv():
    import csv, io
    from flask import Response
    movs, _ = _filtra_mov_material()
    buf = io.StringIO(); buf.write("\ufeff")
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Data/Hora", "Tipo", "Material", "Quantidade", "Saldo após", "Colaborador", "Operador", "Obs"])
    for m in movs:
        w.writerow([m.criado_em.strftime("%d/%m/%Y %H:%M") if m.criado_em else "", m.tipo,
                    m.produto_nome or "", ("%g" % (m.quantidade or 0)), ("%g" % (m.saldo_apos or 0)),
                    m.colaborador_nome or "", (m.operador.nome if m.operador else ""), m.obs or ""])
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=movimentacoes_material.csv"})


@almox_bp.route("/materiais/movimentacoes/pdf")
@_guard("pode_material")
def materiais_mov_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io
    from flask import Response
    movs, ctx = _filtra_mov_material()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Movimentações de material", styles["Title"]),
             Paragraph(f"Período: {ctx['data_ini'] or 'início'} a {ctx['data_fim'] or 'hoje'}", styles["Normal"]),
             Spacer(1, 6)]
    data = [["Data/Hora", "Tipo", "Material", "Qtd", "Saldo após", "Colaborador", "Operador"]]
    for m in movs:
        data.append([m.criado_em.strftime("%d/%m/%y %H:%M") if m.criado_em else "", m.tipo,
                     m.produto_nome or "", ("%g" % (m.quantidade or 0)), ("%g" % (m.saldo_apos or 0)),
                     m.colaborador_nome or "", (m.operador.nome if m.operador else "")])
    if len(data) == 1:
        data.append(["—"] * 7)
    t = Table(data, repeatRows=1, colWidths=[26*mm, 18*mm, 70*mm, 20*mm, 24*mm, 40*mm, 40*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5246")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=movimentacoes_material.pdf"})


@almox_bp.route("/materiais/saldo/csv")
@_guard("pode_material")
def materiais_saldo_csv():
    import csv, io
    from flask import Response
    itens, _ = _filtra_saldo()
    buf = io.StringIO(); buf.write("\ufeff")
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Código", "Material", "Unidade", "Categoria", "Saldo", "Saldo mínimo", "Abaixo do mínimo?"])
    for p in itens:
        w.writerow([p.codigo or "", p.nome, p.unidade, p.categoria or "",
                    ("%g" % (p.saldo or 0)), ("%g" % (p.saldo_minimo or 0)),
                    "SIM" if p.abaixo_minimo else ""])
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=saldo_material.csv"})


@almox_bp.route("/materiais/qr")
@_guard("pode_material")
def materiais_qr():
    ids = request.args.get("ids") or ""
    consulta = ProdutoAlmox.query.filter_by(ativo=True)
    if ids.strip():
        lista_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        consulta = consulta.filter(ProdutoAlmox.id.in_(lista_ids))
    itens = consulta.order_by(ProdutoAlmox.nome).all()
    formato = request.args.get("formato", "a4")
    return render_template("almox/materiais_qr.html", itens=itens, qr_svg=_qr_svg, formato=formato)


# ----- Coletor: saída de MATERIAL (via QR do colaborador + QR do produto) -----
@almox_bp.route("/coletor/api/produto/<path:qr_uid>")
@modulo_required
def coletor_api_produto(qr_uid):
    if not _aba_coletor_ok("material"): return _nega_aba("material")
    uid = _uid_limpo(qr_uid)
    p = ProdutoAlmox.query.filter_by(qr_uid=uid, ativo=True).first()
    if not p:
        return jsonify(ok=False, erro="Material não encontrado."), 404
    return jsonify(ok=True, id=p.id, nome=p.nome, unidade=p.unidade, saldo=p.saldo or 0)


@almox_bp.route("/coletor/api/material-saida", methods=["POST"])
@modulo_required
def coletor_api_material_saida():
    if not _aba_coletor_ok("material"): return _nega_aba("material")
    """payload: {colaborador_id, senha, itens:[{produto_id, qtd}]}"""
    data = request.get_json(silent=True) or {}
    colab = db.session.get(Colaborador, data.get("colaborador_id") or 0)
    if not colab or not colab.ativo:
        return jsonify(ok=False, erro="Colaborador inválido."), 400
    senha = data.get("senha") or ""
    if not colab.tem_senha:
        if len(senha) < 4:
            return jsonify(ok=False, erro="Primeiro uso: defina uma senha de ao menos 4 dígitos."), 400
        colab.set_senha(senha)
    elif not colab.check_senha(senha):
        return jsonify(ok=False, erro="Senha do colaborador inválida."), 403
    itens = data.get("itens") or []
    if not itens:
        return jsonify(ok=False, erro="Nenhum material para confirmar."), 400
    # Valida saldo ANTES de aplicar (prevenção de estoque negativo)
    faltas = []
    plano = []
    for it in itens:
        p = db.session.get(ProdutoAlmox, it.get("produto_id") or 0)
        q = _num(it.get("qtd"))
        if not p or q <= 0:
            continue
        if q > (p.saldo or 0):
            faltas.append(f"{p.nome} (saldo {p.saldo:g}, pedido {q:g})")
        else:
            plano.append((p, q))
    if faltas:
        return jsonify(ok=False, erro="Saldo insuficiente: " + "; ".join(faltas)), 400
    feitas = []
    for p, q in plano:
        p.saldo = (p.saldo or 0) - q
        db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="saida",
                       quantidade=q, saldo_apos=p.saldo, colaborador_id=colab.id,
                       colaborador_nome=colab.nome, operador_id=_op_id(), obs="Coletor"))
        feitas.append(f"{q:g} {p.unidade} de {p.nome}")
    _log("Coletor", f"{colab.nome}: saída de material — " + "; ".join(feitas))
    db.session.commit()
    return jsonify(ok=True, resumo=feitas, colaborador=colab.nome)


# ----- Estoque negativo (detecção/resolução — roadmap §9) -----
@almox_bp.route("/materiais/negativos")
@_guard("pode_material")
def materiais_negativos():
    negativos = (ProdutoAlmox.query.filter(ProdutoAlmox.ativo == True, ProdutoAlmox.saldo < 0)
                 .order_by(ProdutoAlmox.nome).all())
    return render_template("almox/materiais_negativos.html", negativos=negativos)


# ----- Coletor: MOVIMENTAÇÃO de material entre locais -----
@almox_bp.route("/coletor/api/locais")
@modulo_required
def coletor_api_locais():
    locais = LocalAlmox.query.filter_by(ativo=True).order_by(LocalAlmox.nome).all()
    return jsonify(ok=True, locais=[{"id": l.id, "nome": l.nome} for l in locais])


@almox_bp.route("/coletor/api/mover-legado", methods=["POST"])
@modulo_required
def coletor_api_mover_legado():
    if not _aba_coletor_ok("mover"): return _nega_aba("mover")
    data = request.get_json(silent=True) or {}
    p = db.session.get(ProdutoAlmox, data.get("produto_id") or 0)
    destino = db.session.get(LocalAlmox, data.get("local_id") or 0)
    if not p or not p.ativo:
        return jsonify(ok=False, erro="Material inválido."), 400
    if not destino:
        return jsonify(ok=False, erro="Local de destino inválido."), 400
    de = p.local_nome
    p.local_id = destino.id
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="movimentacao",
                   quantidade=0, saldo_apos=p.saldo, local_de=de, local_para=destino.nome,
                   operador_id=_op_id(), obs="Coletor"))
    _log("Coletor", f"{p.nome}: movido de {de} para {destino.nome}")
    db.session.commit()
    return jsonify(ok=True, resumo=[f"{p.nome}: {de} → {destino.nome}"])


# ==================== INVENTÁRIO DE MATERIAL ====================
@almox_bp.route("/materiais/inventario", methods=["GET"])
@_guard("pode_material")
def inventario():
    itens = ProdutoAlmox.query.filter_by(ativo=True).order_by(ProdutoAlmox.nome).all()
    return render_template("almox/inventario.html", itens=itens)


@almox_bp.route("/materiais/inventario", methods=["POST"])
@_guard("pode_material")
def inventario_salvar():
    itens = ProdutoAlmox.query.filter_by(ativo=True).all()
    ajustados = 0
    for p in itens:
        raw = request.form.get(f"contado_{p.id}")
        if raw is None or str(raw).strip() == "":
            continue                     # item não contado nesta rodada: ignora
        contado = _num(raw)
        if contado != (p.saldo or 0):
            dif = contado - (p.saldo or 0)
            p.saldo = contado
            db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="inventario",
                           quantidade=dif, saldo_apos=contado, operador_id=_op_id(),
                           obs="Inventário (conferência)"))
            ajustados += 1
    _log("Material", f"Inventário aplicado: {ajustados} item(ns) ajustado(s)")
    db.session.commit()
    flash(f"Inventário concluído. {ajustados} item(ns) ajustado(s).", "success")
    return redirect(url_for("almox.materiais"))


@almox_bp.route("/coletor/api/inventario", methods=["POST"])
@modulo_required
def coletor_api_inventario():
    if not _aba_coletor_ok("inventario"): return _nega_aba("inventario")
    """payload: {produto_id, contado}"""
    data = request.get_json(silent=True) or {}
    p = db.session.get(ProdutoAlmox, data.get("produto_id") or 0)
    if not p or not p.ativo:
        return jsonify(ok=False, erro="Material inválido."), 400
    contado = _num(data.get("contado"), None if data.get("contado") in (None, "") else 0)
    if contado is None:
        return jsonify(ok=False, erro="Quantidade contada inválida."), 400
    antigo = p.saldo or 0
    dif = contado - antigo
    p.saldo = contado
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="inventario",
                   quantidade=dif, saldo_apos=contado, operador_id=_op_id(), obs="Inventário (coletor)"))
    _log("Coletor", f"Inventário {p.nome}: {antigo:g} → {contado:g}")
    db.session.commit()
    return jsonify(ok=True, resumo=[f"{p.nome}: {antigo:g} → {contado:g}"], nome=p.nome,
                   antigo=antigo, novo=contado)


# ==================== HIERARQUIA FÍSICA: PLANTA / ARMAZÉM / LOCALIZADOR ====================

@almox_bp.route("/plantas", methods=["GET", "POST"])
@_guard("pode_locais")
def plantas():
    from .models import Planta
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip().upper()
        if not nome:
            flash("Informe o nome da planta.", "danger")
        elif Planta.query.filter(db.func.upper(Planta.nome) == nome).first():
            flash("Já existe uma planta com esse nome.", "warning")
        else:
            db.session.add(Planta(nome=nome))
            _log("Planta", f"Planta cadastrada: {nome}")
            db.session.commit()
            flash("Planta cadastrada.", "success")
        return redirect(url_for("almox.plantas"))
    itens = Planta.query.order_by(Planta.nome).all()
    return render_template("almox/plantas.html", itens=itens)


@almox_bp.route("/plantas/<int:pid>", methods=["POST"])
@_guard("pode_locais")
def planta_editar(pid):
    from .models import Planta
    p = db.session.get(Planta, pid) or abort(404)
    p.nome = (request.form.get("nome") or p.nome).strip().upper()
    p.ativo = request.form.get("ativo") == "1"
    db.session.commit()
    flash("Planta atualizada.", "success")
    return redirect(url_for("almox.plantas"))


@almox_bp.route("/armazens", methods=["GET", "POST"])
@_guard("pode_locais")
def armazens():
    from .models import Armazem, Planta
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip().upper()
        planta_id = request.form.get("planta_id") or None
        if not nome or not planta_id:
            flash("Informe o nome do armazém e a planta.", "danger")
        else:
            db.session.add(Armazem(nome=nome, planta_id=int(planta_id)))
            _log("Armazém", f"Armazém cadastrado: {nome}")
            db.session.commit()
            flash("Armazém cadastrado.", "success")
        return redirect(url_for("almox.armazens"))
    itens = Armazem.query.order_by(Armazem.nome).all()
    plantas_l = Planta.query.filter_by(ativo=True).order_by(Planta.nome).all()
    return render_template("almox/armazens.html", itens=itens, plantas=plantas_l)


@almox_bp.route("/armazens/<int:aid>", methods=["POST"])
@_guard("pode_locais")
def armazem_editar(aid):
    from .models import Armazem
    a = db.session.get(Armazem, aid) or abort(404)
    a.nome = (request.form.get("nome") or a.nome).strip().upper()
    if request.form.get("planta_id"):
        a.planta_id = int(request.form.get("planta_id"))
    a.ativo = request.form.get("ativo") == "1"
    db.session.commit()
    flash("Armazém atualizado.", "success")
    return redirect(url_for("almox.armazens"))


@almox_bp.route("/localizadores", methods=["GET", "POST"])
@_guard("pode_locais")
def localizadores():
    import secrets
    from .models import Armazem, Localizador
    if request.method == "POST":
        armazem_id = request.form.get("armazem_id") or None
        fila = (request.form.get("fila") or "").strip().upper()[:1]
        estante = request.form.get("estante")
        nivel = request.form.get("nivel")
        if not (armazem_id and fila.isalpha() and (estante or "").isdigit() and (nivel or "").isdigit()):
            flash("Preencha armazém, fila (1 letra), estante e nível (números).", "danger")
            return redirect(url_for("almox.localizadores"))
        existe = Localizador.query.filter_by(armazem_id=int(armazem_id), fila=fila,
                                             estante=int(estante), nivel=int(nivel)).first()
        if existe:
            flash("Esse localizador já existe.", "warning")
            return redirect(url_for("almox.localizadores"))
        uid = "LOC-" + secrets.token_hex(4).upper()
        db.session.add(Localizador(armazem_id=int(armazem_id), fila=fila,
                                   estante=int(estante), nivel=int(nivel), qr_uid=uid))
        _log("Localizador", f"Localizador {fila}*{estante}*{nivel} cadastrado")
        db.session.commit()
        flash("Localizador cadastrado.", "success")
        return redirect(url_for("almox.localizadores"))
    itens = Localizador.query.order_by(Localizador.armazem_id, Localizador.fila,
                                       Localizador.estante, Localizador.nivel).all()
    armazens_l = Armazem.query.filter_by(ativo=True).order_by(Armazem.nome).all()
    return render_template("almox/localizadores.html", itens=itens, armazens=armazens_l)


@almox_bp.route("/localizadores/<int:lid>/toggle", methods=["POST"])
@_guard("pode_locais")
def localizador_toggle(lid):
    from .models import Localizador
    l = db.session.get(Localizador, lid) or abort(404)
    l.ativo = not l.ativo
    db.session.commit()
    flash("Localizador atualizado.", "success")
    return redirect(url_for("almox.localizadores"))


@almox_bp.route("/localizadores/gerar", methods=["GET", "POST"])
@_guard("pode_locais")
def localizadores_gerar():
    import secrets
    from .models import Armazem, Localizador
    if request.method == "POST":
        armazem_id = request.form.get("armazem_id") or None
        f_ini = (request.form.get("fila_ini") or "").strip().upper()[:1]
        f_fim = (request.form.get("fila_fim") or "").strip().upper()[:1]
        try:
            e_ini = int(request.form.get("estante_ini")); e_fim = int(request.form.get("estante_fim"))
            n_ini = int(request.form.get("nivel_ini")); n_fim = int(request.form.get("nivel_fim"))
        except (TypeError, ValueError):
            flash("Estante e nível devem ser números.", "danger")
            return redirect(url_for("almox.localizadores_gerar"))
        if not (armazem_id and f_ini.isalpha() and f_fim.isalpha() and f_ini <= f_fim
                and e_ini <= e_fim and n_ini <= n_fim):
            flash("Confira os intervalos (fila A→Z, estante e nível crescentes).", "danger")
            return redirect(url_for("almox.localizadores_gerar"))
        criados, pulados = 0, 0
        existentes = {(l.fila, l.estante, l.nivel) for l in
                      Localizador.query.filter_by(armazem_id=int(armazem_id)).all()}
        for cod in range(ord(f_ini), ord(f_fim) + 1):
            fila = chr(cod)
            for est in range(e_ini, e_fim + 1):
                for niv in range(n_ini, n_fim + 1):
                    if (fila, est, niv) in existentes:
                        pulados += 1
                        continue
                    uid = "LOC-" + secrets.token_hex(4).upper()
                    db.session.add(Localizador(armazem_id=int(armazem_id), fila=fila,
                                               estante=est, nivel=niv, qr_uid=uid))
                    criados += 1
        _log("Localizador", f"Gerados {criados} localizadores em massa (pulados {pulados})")
        db.session.commit()
        flash(f"{criados} localizador(es) gerado(s). {pulados} já existiam.", "success")
        return redirect(url_for("almox.localizadores"))
    armazens_l = Armazem.query.filter_by(ativo=True).order_by(Armazem.nome).all()
    return render_template("almox/localizadores_gerar.html", armazens=armazens_l)


@almox_bp.route("/relatorios")
@_guard("pode_relatorios")
def relatorios_central():
    """Central de relatórios: reúne num lugar só todos os relatórios e exportações."""
    return render_template("almox/relatorios_central.html")


@almox_bp.route("/dashboard")
@_guard("pode_relatorios")
def dashboard():
    from datetime import timedelta
    # ---- Material ----
    ativos = ProdutoAlmox.query.filter_by(ativo=True).all()
    mat_itens = len(ativos)
    mat_baixo = sum(1 for p in ativos if p.abaixo_minimo)
    mat_unidades = int(sum(p.saldo or 0 for p in ativos))
    baixas_pend = AjusteInventario.query.filter_by(status="pendente").count()
    nfs_classificar = NotaFiscalAlmox.query.filter(NotaFiscalAlmox.classificacao.is_(None)).count()

    # ---- Solicitações por status (dinâmico) ----
    LABELS = {
        "AGUARDANDO_APROVACAO": ("Aguardando aprovação", "#f0be3c"),
        "AGUARDANDO_ENVIO_COTACAO": ("Aguardando envio p/ cotação", "#5b9bd5"),
        "AGUARDANDO_RECEBIMENTO_COTACAO": ("Aguardando cotação", "#5b9bd5"),
        "AGUARDANDO_DEFINICAO_FORNECEDOR": ("Definição de fornecedor", "#8a7bd5"),
        "AGUARDANDO_CHEGADA": ("Aguardando chegada", "#8a7bd5"),
        "CONCLUIDO": ("Concluídas", "#32CAA0"),
        "CONCLUIDA": ("Concluídas", "#32CAA0"),
        "CANCELADA": ("Canceladas", "#ff8b82"),
        "CANCELADO": ("Canceladas", "#ff8b82"),
    }
    rows = db.session.query(Solicitacao.status, db.func.count(Solicitacao.id)).group_by(Solicitacao.status).all()
    smax = max([c for _, c in rows], default=0) or 1
    solic = []
    for st, c in sorted(rows, key=lambda r: -r[1]):
        lbl, cor = LABELS.get(st, (st or "—", "#9a9a9a"))
        solic.append({"status": st, "label": lbl, "cor": cor, "n": c, "pct": round(c / smax * 100)})

    # ---- Extintores por situação (donut) ----
    no_prazo = venc = prox = 0
    for e in Extintor.query.filter_by(ativo=True).all():
        k = _situacao_extintor(e)[0]
        if k == "NO_PRAZO":
            no_prazo += 1
        elif k == "PROX_VENC":
            prox += 1
        else:
            venc += 1
    ext_total = no_prazo + venc + prox
    segs, cum = [], 0.0
    for val, cor, key, lbl in [(no_prazo, "#32CAA0", "NO_PRAZO", "No prazo"),
                               (prox, "#f0be3c", "PROX_VENC", "Vencendo"),
                               (venc, "#ff8b82", "VENCIDO", "Vencido / irregular")]:
        ln = (val / ext_total * 100) if ext_total else 0
        segs.append({"cor": cor, "dash": f"{ln:.1f} {100 - ln:.1f}",
                     "offset": f"{(25 - cum):.1f}", "key": key, "label": lbl, "n": val})
        cum += ln

    # ---- Chaves ----
    ch_total = Chave.query.filter_by(ativo=True).count()
    ch_uso = Chave.query.filter_by(ativo=True, status="Em uso").count()

    # ---- Movimentações (7 dias) ----
    desde = datetime.utcnow() - timedelta(days=7)
    mov_ent = mov_sai = 0
    try:
        mov_ent = MovimentacaoMaterial.query.filter(MovimentacaoMaterial.tipo == "entrada",
                                                    MovimentacaoMaterial.criado_em >= desde).count()
        mov_sai = MovimentacaoMaterial.query.filter(MovimentacaoMaterial.tipo == "saida",
                                                    MovimentacaoMaterial.criado_em >= desde).count()
    except Exception:
        pass

    return render_template("almox/dashboard.html",
                           mat_itens=mat_itens, mat_baixo=mat_baixo, mat_unidades=mat_unidades,
                           baixas_pend=baixas_pend, nfs_classificar=nfs_classificar,
                           solic=solic, ext=dict(no_prazo=no_prazo, prox=prox, venc=venc,
                           total=ext_total, segs=segs),
                           ch_total=ch_total, ch_uso=ch_uso, mov_ent=mov_ent, mov_sai=mov_sai,
                           is_admin=_efetivo("is_admin"))


# ==================== IMPORTAÇÃO EM LOTE — COLABORADORES ====================
import csv as _csv
import io as _io

COLAB_CSV_COLS = ["nome", "cpf", "email", "empresa", "cargo", "perfil"]


def _gerar_xlsx(colunas, exemplos, nome_aba="Modelo"):
    """Gera um .xlsx (bytes) com cabeçalho em negrito + linhas de exemplo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = nome_aba
    ws.append(colunas)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ex in exemplos:
        ws.append(ex)
    for i, _col in enumerate(colunas, start=1):
        ws.column_dimensions[chr(64 + i)].width = 22
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def _ler_planilha(arquivo):
    """Lê .xlsx ou .csv e devolve (lista_de_dicts, erro). Cabeçalhos em minúsculo."""
    nome = (arquivo.filename or "").lower()
    dados = arquivo.read()
    if nome.endswith(".xlsx"):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(_io.BytesIO(dados), read_only=True, data_only=True)
        except Exception:
            return None, "Não consegui ler o Excel. Verifique o arquivo."
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            return [], None
        cab = [str(h).strip().lower() if h is not None else "" for h in linhas[0]]
        out = []
        for row in linhas[1:]:
            d = {}
            for i, h in enumerate(cab):
                v = row[i] if i < len(row) else None
                d[h] = "" if v is None else str(v).strip()
            out.append(d)
        return out, None
    # CSV
    try:
        txt = dados.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            txt = dados.decode("latin-1")
        except Exception:
            return None, "Não consegui ler o arquivo. Salve como .xlsx ou CSV UTF-8."
    sep = ";" if txt.count(";") >= txt.count(",") else ","
    leitor = _csv.DictReader(_io.StringIO(txt), delimiter=sep)
    cab = [(h or "").strip().lower() for h in (leitor.fieldnames or [])]
    leitor.fieldnames = cab
    return [ {k: (v or "").strip() for k, v in row.items()} for row in leitor ], None


@almox_bp.route("/colaboradores/modelo.xlsx")
@_guard("pode_colaboradores")
def colaboradores_modelo_csv():
    from flask import Response
    dados = _gerar_xlsx(COLAB_CSV_COLS, [
        ["JOAO DA SILVA", "12345678901", "joao@empresa.com", "OMEGA ENERGIA", "TECNICO", "ALMOXARIFADO"],
        ["MARIA SOUZA", "98765432100", "", "PRESTADORA XYZ", "AUXILIAR", "SOLICITANTE"],
    ], "Colaboradores")
    return Response(dados, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=modelo_colaboradores.xlsx"})


@almox_bp.route("/colaboradores/importar", methods=["POST"])
@_guard("pode_colaboradores")
def colaboradores_importar():
    import secrets
    arq = request.files.get("arquivo")
    if not arq or not arq.filename:
        flash("Selecione um arquivo .xlsx ou CSV.", "danger")
        return redirect(url_for("almox.colaboradores"))
    linhas, erro = _ler_planilha(arq)
    if erro:
        flash(erro, "danger")
        return redirect(url_for("almox.colaboradores"))
    nomes_perfil = {p.nome.upper() for p in PapelColaborador.query.all()}
    existentes = set()
    for c0 in Colaborador.query.filter(Colaborador.ativo.is_(True)).all():
        existentes.add("".join(ch for ch in (c0.cpf or "") if ch.isdigit()))
    criados, pulados, motivos = 0, 0, []
    for i, linha in enumerate(linhas, start=2):
        nome = (linha.get("nome") or "").strip().upper()
        cpf = "".join(ch for ch in (linha.get("cpf") or "") if ch.isdigit())
        email = (linha.get("email") or "").strip().lower()
        empresa = (linha.get("empresa") or "").strip().upper()
        cargo = (linha.get("cargo") or "").strip().upper()
        perfil = (linha.get("perfil") or "").strip().upper()
        if not nome or not cpf:
            pulados += 1; motivos.append(f"linha {i}: nome e CPF são obrigatórios"); continue
        if cpf in existentes:
            pulados += 1; motivos.append(f"linha {i}: CPF {cpf} já cadastrado"); continue
        papel = "COLABORADOR DIVERSO"
        if perfil:
            if perfil in nomes_perfil:
                papel = perfil
            else:
                motivos.append(f"linha {i}: perfil '{perfil}' não existe — entrou como COLABORADOR DIVERSO")
        uid = "COL-" + secrets.token_hex(4).upper()
        while Colaborador.query.filter_by(qr_uid=uid).first():
            uid = "COL-" + secrets.token_hex(4).upper()
        db.session.add(Colaborador(nome=nome, cpf=cpf, email=(email or None),
                                   empresa=empresa, cargo=cargo, papel=papel, qr_uid=uid))
        existentes.add(cpf); criados += 1
    db.session.commit()
    _log("Colaborador", f"Importação em lote: {criados} criado(s), {pulados} pulado(s)")
    resumo = f"Importação concluída: {criados} colaborador(es) criado(s), {pulados} pulado(s)."
    if motivos:
        resumo += " Detalhes: " + "; ".join(motivos[:15]) + ("..." if len(motivos) > 15 else "")
    flash(resumo, "success" if criados else "warning")
    return redirect(url_for("almox.colaboradores"))


# ==================== CADASTRO: FABRICANTES ====================
@almox_bp.route("/fabricantes")
@_guard("pode_material")
def fabricantes():
    mostrar_inativos = request.args.get("inativos") == "1"
    q = Fabricante.query
    if not mostrar_inativos:
        q = q.filter_by(ativo=True)
    itens = q.order_by(Fabricante.nome).all()
    return render_template("almox/fabricantes.html", itens=itens, mostrar_inativos=mostrar_inativos)


@almox_bp.route("/fabricantes/novo", methods=["POST"])
@_guard("pode_material")
def fabricante_novo():
    nome = (request.form.get("nome") or "").strip().upper()
    if not nome:
        flash("Informe o nome do fabricante.", "danger")
        return redirect(url_for("almox.fabricantes"))
    if Fabricante.query.filter(db.func.upper(Fabricante.nome) == nome).first():
        flash("Já existe um fabricante com esse nome.", "warning")
        return redirect(url_for("almox.fabricantes"))
    db.session.add(Fabricante(nome=nome, ativo=True))
    _log("Material", f"Fabricante cadastrado: {nome}")
    db.session.commit()
    flash("Fabricante cadastrado.", "success")
    return redirect(url_for("almox.fabricantes"))


@almox_bp.route("/fabricantes/<int:fid>/editar", methods=["POST"])
@_guard("pode_material")
def fabricante_editar(fid):
    f = Fabricante.query.get_or_404(fid)
    nome = (request.form.get("nome") or "").strip().upper()
    if nome and nome != (f.nome or "").upper():
        existe = Fabricante.query.filter(db.func.upper(Fabricante.nome) == nome, Fabricante.id != fid).first()
        if existe:
            flash("Já existe um fabricante com esse nome.", "warning")
            return redirect(url_for("almox.fabricantes"))
        f.nome = nome
        db.session.commit()
        flash("Fabricante atualizado.", "success")
    return redirect(url_for("almox.fabricantes"))


@almox_bp.route("/fabricantes/<int:fid>/toggle", methods=["POST"])
@_guard("pode_material")
def fabricante_toggle(fid):
    f = Fabricante.query.get_or_404(fid)
    f.ativo = not f.ativo
    db.session.commit()
    flash("Fabricante " + ("reativado." if f.ativo else "desativado."), "success")
    return redirect(url_for("almox.fabricantes"))


# ==================== COLETOR REFORMADO — BLOCOS 2 e 3 ====================
import unicodedata as _ud

def _norm(s):
    s = (s or "").strip().lower()
    return "".join(c for c in _ud.normalize("NFKD", s) if not _ud.combining(c))


def _resolver_localizador(termo):
    """Aceita qr_uid do localizador OU o código no formato A*1*3."""
    from .models import Localizador
    t = _uid_limpo(termo).strip()
    loc = Localizador.query.filter_by(qr_uid=t, ativo=True).first()
    if loc:
        return loc
    if "*" in t:
        partes = t.upper().split("*")
        if len(partes) == 3 and partes[1].isdigit() and partes[2].isdigit():
            return (Localizador.query.filter_by(fila=partes[0], estante=int(partes[1]),
                    nivel=int(partes[2]), ativo=True).first())
    return None


@almox_bp.route("/coletor/api/localizador/<path:qr_uid>")
@modulo_required
def coletor_api_localizador(qr_uid):
    if not _aba_material_qualquer(): return _nega_aba("material", "Seu perfil nao usa o coletor para material/localizadores.")
    loc = _resolver_localizador(qr_uid)
    if not loc:
        return jsonify(ok=False, erro="Localizador não encontrado."), 404
    return jsonify(ok=True, id=loc.id, codigo=loc.codigo, caminho=loc.caminho)


@almox_bp.route("/coletor/api/item/<path:qr_uid>")
@modulo_required
def coletor_api_item(qr_uid):
    """Unificado: reconhece pelo QR se é CHAVE (CH-/QUAD-) ou MATERIAL."""
    uid = _uid_limpo(qr_uid)
    up = uid.upper()
    if up.startswith("CH-") or up.startswith("QUAD-"):
        if not _aba_coletor_ok("chaves"):
            return _nega_aba("chaves", "Seu perfil não usa o coletor para chaves.")
        c = Chave.query.filter_by(qr_uid=uid, ativo=True).first()
        if not c:
            return jsonify(ok=False, erro="Chave não encontrada."), 404
        return jsonify(ok=True, tipo="chave", id=c.id, nome=c.descricao,
                       info=c.quadro_nome, status=c.status, com_quem=c.com_quem)
    p = ProdutoAlmox.query.filter_by(qr_uid=uid, ativo=True).first()
    if p:
        if not _aba_material_qualquer():
            return _nega_aba("material", "Seu perfil não usa o coletor para material.")
        return jsonify(ok=True, tipo="material", id=p.id, nome=p.nome,
                       info=(p.unidade or "UN"), saldo=p.saldo or 0, localizador=p.local_nome)
    return jsonify(ok=False, erro="Item não encontrado."), 404


@almox_bp.route("/coletor/api/buscar")
@modulo_required
def coletor_api_buscar():
    """Busca pesquisável para os campos do coletor. tipo=colab|loc|item|fabricante|nf; q=texto."""
    from .models import Localizador
    tipo = request.args.get("tipo", "")
    q = _norm(request.args.get("q", ""))
    res = []
    # bloqueios por aba
    if tipo == "loc" and not _aba_material_qualquer():
        return _nega_aba("material")
    if tipo in ("fabricante", "nf") and not _aba_coletor_ok("material"):
        return _nega_aba("material")
    if tipo == "colab" and not (_aba_coletor_ok("chaves") or _aba_coletor_ok("material")):
        return _nega_aba("chaves")
    if tipo == "colab":
        for c in Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome).all():
            if not q or q in _norm(c.nome) or q in _norm(c.empresa):
                res.append({"id": c.id, "nome": c.nome, "info": c.empresa or ""})
    elif tipo == "loc":
        for l in Localizador.query.filter_by(ativo=True).all():
            if not q or q in _norm(l.codigo) or q in _norm(l.caminho):
                res.append({"id": l.id, "nome": l.codigo, "info": l.caminho})
        res.sort(key=lambda x: x["nome"])
    elif tipo == "item":
        if _aba_material_qualquer():
            for p in ProdutoAlmox.query.filter_by(ativo=True).order_by(ProdutoAlmox.nome).all():
                if not q or q in _norm(p.nome) or q in _norm(p.codigo) or q in _norm(p.codigo_barras):
                    res.append({"id": p.id, "tipo": "material", "nome": p.nome,
                                "info": "Material · " + (p.local_nome or "—"), "saldo": p.saldo or 0})
        if _aba_coletor_ok("chaves"):
            for c in Chave.query.filter_by(ativo=True).order_by(Chave.descricao).all():
                if not q or q in _norm(c.descricao) or q in _norm(c.quadro_nome):
                    res.append({"id": c.id, "tipo": "chave", "nome": c.descricao,
                                "info": "Chave · " + (c.quadro_nome or "—"), "status": c.status})
    elif tipo == "fabricante":
        for f in Fabricante.query.filter_by(ativo=True).order_by(Fabricante.nome).all():
            if not q or q in _norm(f.nome):
                res.append({"id": f.id, "nome": f.nome, "info": "fabricante"})
    elif tipo == "nf":
        for n in NotaFiscalAlmox.query.order_by(NotaFiscalAlmox.criado_em.desc()).all():
            if not q or q in _norm(n.numero) or q in _norm(n.fornecedor_nome):
                res.append({"id": n.id, "nome": n.rotulo,
                            "info": (("R$ %.2f" % n.valor) if n.valor else "")})
    return jsonify(ok=True, itens=res[:40])


def _valida_senha_colab(colab, senha):
    if not colab.tem_senha:
        if len(senha or "") < 4:
            return "Primeiro uso: defina uma senha de ao menos 4 dígitos."
        colab.set_senha(senha)
        return None
    if not colab.check_senha(senha):
        return "Senha do colaborador inválida."
    return None


@almox_bp.route("/coletor/api/retirar", methods=["POST"])
@modulo_required
def coletor_api_retirar():
    """RETIRADA unificada (chave/material) com localizador por item.
    payload: {colaborador_id, senha, itens:[{tipo,id,qtd,localizador_id,localizador_cod}]}"""
    data = request.get_json(silent=True) or {}
    colab = db.session.get(Colaborador, data.get("colaborador_id") or 0)
    if not colab or not colab.ativo:
        return jsonify(ok=False, erro="Colaborador inválido."), 400
    err = _valida_senha_colab(colab, data.get("senha"))
    if err:
        return jsonify(ok=False, erro=err), 403
    itens = data.get("itens") or []
    if not itens:
        return jsonify(ok=False, erro="Cesta vazia."), 400
    _tipos = {it.get("tipo") for it in itens}
    if "chave" in _tipos and not _aba_coletor_ok("chaves"):
        return _nega_aba("chaves")
    if "material" in _tipos and not _aba_coletor_ok("material"):
        return _nega_aba("material")
    plano_mat, faltas, plano_chave = [], [], []
    for it in itens:
        if it.get("tipo") == "material":
            p = db.session.get(ProdutoAlmox, it.get("id") or 0)
            qn = _num(it.get("qtd"), 1)
            if not p or qn <= 0:
                continue
            loc_id = it.get("localizador_id")
            loccod = it.get("localizador_cod") or p.local_nome
            if loc_id:
                linha = p.estoque_em(int(loc_id))
                disp = (linha.quantidade or 0) if linha else 0
                if qn > disp:
                    faltas.append(f"{p.nome} em {loccod} (tem {disp:g}, pediu {qn:g})")
                else:
                    plano_mat.append((p, qn, int(loc_id), loccod))
            else:
                if qn > (p.saldo or 0):
                    faltas.append(f"{p.nome} (saldo {p.saldo:g}, pedido {qn:g})")
                else:
                    plano_mat.append((p, qn, None, loccod))
        elif it.get("tipo") == "chave":
            c = db.session.get(Chave, it.get("id") or 0)
            if c and c.ativo and c.status == "Disponível":
                plano_chave.append(c)
    if faltas:
        return jsonify(ok=False, erro="Saldo insuficiente: " + "; ".join(faltas)), 400
    feitas = []
    for p, qn, loc_id, loccod in plano_mat:
        if loc_id:
            p.ajustar_estoque(loc_id, -qn)
        else:
            p.saldo = (p.saldo or 0) - qn
        db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="saida",
                       quantidade=qn, saldo_apos=p.saldo, local_de=loccod, colaborador_id=colab.id,
                       colaborador_nome=colab.nome, operador_id=_op_id(), obs="Coletor/retirada"))
        feitas.append(f"{qn:g} {p.unidade} de {p.nome} ({loccod})")
    for c in plano_chave:
        c.status = "Em uso"; c.com_quem = colab.nome
        db.session.add(MovimentacaoChave(chave_id=c.id, chave_desc=c.descricao, quadro_nome=c.quadro_nome,
                       colaborador_id=colab.id, colaborador_nome=colab.nome, acao="retirada", operador_id=_op_id()))
        feitas.append(f"Chave {c.descricao}")
    _log("Coletor", f"{colab.nome}: retirada — " + "; ".join(feitas))
    db.session.commit()
    return jsonify(ok=True, resumo=feitas, colaborador=colab.nome)


@almox_bp.route("/coletor/api/devolver", methods=["POST"])
@modulo_required
def coletor_api_devolver():
    """DEVOLUÇÃO unificada (sem localizador, por decisão).
    payload: {colaborador_id, senha, itens:[{tipo,id,qtd}]}"""
    data = request.get_json(silent=True) or {}
    colab = db.session.get(Colaborador, data.get("colaborador_id") or 0)
    if not colab or not colab.ativo:
        return jsonify(ok=False, erro="Colaborador inválido."), 400
    err = _valida_senha_colab(colab, data.get("senha"))
    if err:
        return jsonify(ok=False, erro=err), 403
    itens = data.get("itens") or []
    if not itens:
        return jsonify(ok=False, erro="Cesta vazia."), 400
    _tipos = {it.get("tipo") for it in itens}
    if "chave" in _tipos and not _aba_coletor_ok("chaves"):
        return _nega_aba("chaves")
    if "material" in _tipos and not _aba_coletor_ok("material"):
        return _nega_aba("material")
    feitas = []
    for it in itens:
        if it.get("tipo") == "material":
            p = db.session.get(ProdutoAlmox, it.get("id") or 0)
            qn = _num(it.get("qtd"), 1)
            if not p or qn <= 0:
                continue
            p.ajustar_estoque(None, qn)
            db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="entrada",
                           quantidade=qn, saldo_apos=p.saldo, local_de="não atribuído", colaborador_id=colab.id,
                           colaborador_nome=colab.nome, operador_id=_op_id(), obs="Coletor/devolução"))
            feitas.append(f"{qn:g} {p.unidade} de {p.nome}")
        elif it.get("tipo") == "chave":
            c = db.session.get(Chave, it.get("id") or 0)
            if c and c.ativo and c.status == "Em uso":
                retirou = c.com_quem or "—"
                c.status = "Disponível"; c.com_quem = None
                db.session.add(MovimentacaoChave(chave_id=c.id, chave_desc=c.descricao, quadro_nome=c.quadro_nome,
                               colaborador_id=colab.id, colaborador_nome=colab.nome, retirado_por=retirou,
                               acao="devolucao", operador_id=_op_id()))
                feitas.append(f"Chave {c.descricao}")
    _log("Coletor", f"{colab.nome}: devolução — " + "; ".join(feitas))
    db.session.commit()
    return jsonify(ok=True, resumo=feitas, colaborador=colab.nome)


@almox_bp.route("/coletor/api/entrada", methods=["POST"])
@modulo_required
def coletor_api_entrada():
    """ENTRADA de item comprado. payload:
    {produto_id, qtd, valor_unit, fabricante_id, fabricante_nd,
     opcionais:{tag,ca,validade,validade_calib,lote},
     nf:{modo:'sem'|'vincular'|'manual', nf_id, fornecedor, numero}}"""
    if not _aba_coletor_ok("material"):
        return _nega_aba("material")
    from datetime import datetime as _dt
    data = request.get_json(silent=True) or {}
    p = db.session.get(ProdutoAlmox, data.get("produto_id") or 0)
    if not p or not p.ativo:
        return jsonify(ok=False, erro="Item inválido. Cadastre-o antes (fica pendente de aprovação)."), 400
    if p.pendente_aprovacao:
        return jsonify(ok=False, erro="Este item está pendente de aprovação do admin."), 400
    qn = _num(data.get("qtd"), 0)
    if qn <= 0:
        return jsonify(ok=False, erro="Informe a quantidade."), 400
    valor = _num(data.get("valor_unit"), 0)
    # fabricante (último usado fica no item)
    if not data.get("fabricante_nd") and data.get("fabricante_id"):
        p.fabricante_id = int(data["fabricante_id"])
    # nota fiscal
    nf = data.get("nf") or {}
    modo = nf.get("modo", "sem")
    nf_obj = None
    notifs = []
    if modo == "vincular" and nf.get("nf_id"):
        nf_obj = db.session.get(NotaFiscalAlmox, int(nf["nf_id"]))
    elif modo == "manual" and (nf.get("numero") or nf.get("fornecedor")):
        nf_obj = NotaFiscalAlmox(numero=(nf.get("numero") or "").strip(),
                                 fornecedor_nome=(nf.get("fornecedor") or "").strip().upper(),
                                 origem="manual")
        db.session.add(nf_obj); db.session.flush()
        notifs.append(NotificacaoAlmox(tipo="nf_sem_cadastro", titulo="NF recebida sem cadastro prévio",
                      texto=f"Entrada de {p.nome}: NF {nf_obj.numero} ({nf_obj.fornecedor_nome}) informada manualmente.",
                      ref_id=nf_obj.id))
    if nf_obj is not None and not nf_obj.classificacao:
        notifs.append(NotificacaoAlmox(tipo="classificar_nf", titulo="Classificar OPEX/CAPEX",
                      texto=f"Nota {nf_obj.rotulo} vinculada à entrada de {p.nome}. Classifique OPEX ou CAPEX.",
                      ref_id=nf_obj.id))
    # aplica entrada (vai para "não atribuído" até ser movido a uma prateleira)
    p.ajustar_estoque(None, qn)
    opc = data.get("opcionais") or {}
    resumo_opc = ", ".join(f"{k}={v}" for k, v in opc.items() if v)
    obs = "Coletor/entrada"
    if valor:
        obs += f" · vlr un R$ {valor:g}"
    if nf_obj is not None:
        obs += f" · {nf_obj.rotulo}"
    if resumo_opc:
        obs += f" · {resumo_opc}"
    db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="entrada",
                   quantidade=qn, saldo_apos=p.saldo, operador_id=_op_id(), obs=obs))
    for n in notifs:
        db.session.add(n)
    _log("Coletor", f"Entrada: {qn:g} de {p.nome}" + (f" · {nf_obj.rotulo}" if nf_obj else ""))
    db.session.commit()
    return jsonify(ok=True, resumo=[f"Entrada de {qn:g} {p.unidade} em {p.nome} (saldo {p.saldo:g})"],
                   notificou=len(notifs) > 0)


@almox_bp.route("/coletor/api/item-pendente", methods=["POST"])
@modulo_required
def coletor_api_item_pendente():
    if not _aba_coletor_ok("material"): return _nega_aba("material")
    """Cadastro rápido de item NÃO existente durante a entrada: cria PENDENTE de aprovação."""
    import secrets
    data = request.get_json(silent=True) or {}
    nome = (data.get("nome") or "").strip().upper()
    if not nome:
        return jsonify(ok=False, erro="Informe o nome do item."), 400
    uid = "MAT-" + secrets.token_hex(4).upper()
    while ProdutoAlmox.query.filter_by(qr_uid=uid).first():
        uid = "MAT-" + secrets.token_hex(4).upper()
    p = ProdutoAlmox(nome=nome, codigo_barras=(data.get("codigo_barras") or "").strip() or None,
                     unidade=(data.get("unidade") or "UN").strip().upper(), saldo=0,
                     qr_uid=uid, pendente_aprovacao=True, ativo=True)
    db.session.add(p); db.session.flush()
    db.session.add(NotificacaoAlmox(tipo="item_pendente", titulo="Item novo aguardando aprovação",
                   texto=f"Item '{nome}' criado na entrada do coletor. Aprove para disponibilizar.", ref_id=p.id))
    _log("Coletor", f"Item pendente criado na entrada: {nome}")
    db.session.commit()
    return jsonify(ok=True, id=p.id, nome=p.nome, pendente=True)


@almox_bp.route("/coletor/api/item-opcionais/<int:pid>")
@modulo_required
def coletor_api_item_opcionais(pid):
    if not _aba_coletor_ok("material"): return _nega_aba("material")
    p = db.session.get(ProdutoAlmox, pid)
    if not p or not p.ativo:
        return jsonify(ok=False, erro="Item não encontrado."), 404
    fab = p.fabricante
    return jsonify(ok=True, opcionais=p.opcionais_ativos,
                   fabricante_id=p.fabricante_id, fabricante_nome=(fab.nome if fab else None))


# ==================== COLETOR — BLOCO 4: INVENTÁRIO ====================
@almox_bp.route("/coletor/api/localizador-itens/<int:loc_id>")
@modulo_required
def coletor_api_localizador_itens(loc_id):
    if not _aba_material_qualquer(): return _nega_aba("material")
    """Itens estocados no localizador (saldo POR localizador), para conferência de inventário."""
    linhas = EstoqueLocalizador.query.filter_by(localizador_id=loc_id).all()
    itens = []
    for l in linhas:
        p = l.produto
        if not p or not p.ativo:
            continue
        itens.append({"id": p.id, "nome": p.nome, "saldo": l.quantidade or 0,
                      "unidade": p.unidade or "UN", "qr_uid": p.qr_uid})
    itens.sort(key=lambda x: x["nome"])
    return jsonify(ok=True, itens=itens)


def _current_pode_conferir(senha):
    """Valida a senha do operador logado (almoxarife) para confirmar inventário/ajustes."""
    u = current_user
    try:
        return bool(senha) and u.check_senha(senha)
    except Exception:
        return False


@almox_bp.route("/coletor/api/inventario-salvar", methods=["POST"])
@modulo_required
def coletor_api_inventario_salvar():
    if not _aba_coletor_ok("inventario"): return _nega_aba("inventario")
    """payload: {localizador_id, localizador_cod, senha, itens:[{produto_id, contado}]}
    Acréscimo aplica na hora; BAIXA fica pendente de aprovação do admin (não altera saldo ainda)."""
    data = request.get_json(silent=True) or {}
    if not _current_pode_conferir(data.get("senha")):
        return jsonify(ok=False, erro="Senha do almoxarife inválida."), 403
    loc_cod = data.get("localizador_cod") or ""
    loc_id = data.get("localizador_id")
    itens = data.get("itens") or []
    aplicados, pendentes = 0, 0
    for it in itens:
        p = db.session.get(ProdutoAlmox, it.get("produto_id") or 0)
        if not p or not p.ativo:
            continue
        raw = it.get("contado")
        if raw is None or str(raw).strip() == "":
            continue
        contado = _num(raw, None)
        if contado is None:
            continue
        # saldo ATUAL naquele localizador
        linha = p.estoque_em(int(loc_id)) if loc_id else None
        atual = (linha.quantidade or 0) if linha else (p.saldo or 0)
        if contado == atual:
            continue
        dif = contado - atual
        tipo = "acrescimo" if dif > 0 else "baixa"
        aj = AjusteInventario(produto_id=p.id, produto_nome=p.nome, localizador_cod=loc_cod,
                              localizador_id=int(loc_id) if loc_id else None,
                              saldo_antes=atual, saldo_novo=contado, diferenca=dif, tipo=tipo,
                              operador_id=_op_id(), operador_nome=getattr(current_user, "nome", None))
        if tipo == "acrescimo":
            if loc_id:
                p.ajustar_estoque(int(loc_id), dif)
            else:
                p.saldo = contado
            aj.status = "aplicado"
            db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="inventario",
                           quantidade=dif, saldo_apos=p.saldo, local_de=loc_cod, operador_id=_op_id(),
                           obs="Inventário (acréscimo)"))
            aplicados += 1
        else:
            aj.status = "pendente"        # baixa NÃO altera saldo até o admin aprovar
            db.session.add(NotificacaoAlmox(tipo="inventario_baixa", titulo="Baixa de inventário a aprovar",
                           texto=f"{p.nome} em {loc_cod}: {atual:g} → {contado:g} (baixa de {abs(dif):g}). Aprove para aplicar.",
                           ref_id=p.id))
            pendentes += 1
        db.session.add(aj)
    _log("Material", f"Inventário {loc_cod}: {aplicados} aplicado(s), {pendentes} baixa(s) pendente(s)")
    db.session.commit()
    return jsonify(ok=True, aplicados=aplicados, pendentes=pendentes,
                   resumo=(f"{aplicados} ajuste(s) aplicado(s); {pendentes} baixa(s) aguardando aprovação do admin."
                           if pendentes else f"{aplicados} ajuste(s) aplicado(s)."))


# ----- Baixas pendentes: aprovação (admin) -----
@almox_bp.route("/inventario/pendentes")
@_guard("is_admin")
def inventario_pendentes():
    pend = AjusteInventario.query.filter_by(status="pendente").order_by(AjusteInventario.criado_em.desc()).all()
    return render_template("almox/inventario_pendentes.html", pendentes=pend)


@almox_bp.route("/inventario/pendentes/<int:aid>/<acao>", methods=["POST"])
@_guard("is_admin")
def inventario_pendente_decidir(aid, acao):
    aj = AjusteInventario.query.get_or_404(aid)
    if aj.status != "pendente":
        flash("Este ajuste já foi decidido.", "warning")
        return redirect(url_for("almox.inventario_pendentes"))
    if acao == "aprovar":
        p = db.session.get(ProdutoAlmox, aj.produto_id)
        if p:
            if aj.localizador_id:
                p.ajustar_estoque(aj.localizador_id, aj.diferenca)   # diferenca é negativa (baixa)
            else:
                p.saldo = aj.saldo_novo
            db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="inventario",
                           quantidade=aj.diferenca, saldo_apos=p.saldo, local_de=aj.localizador_cod,
                           operador_id=_op_id(), obs="Inventário (baixa aprovada)"))
        aj.status = "aplicado"
        flash("Baixa aprovada e aplicada ao saldo.", "success")
    else:
        aj.status = "reprovado"
        flash("Baixa reprovada. Saldo mantido.", "success")
    aj.decidido_por = getattr(current_user, "nome", None)
    aj.decidido_em = datetime.utcnow()
    db.session.commit()
    return redirect(url_for("almox.inventario_pendentes"))


# ----- Relatório de PERDAS (baixas de inventário) por período -----
@almox_bp.route("/relatorios/perdas")
@_guard("pode_material")
def relatorio_perdas():
    de = request.args.get("de") or ""
    ate = request.args.get("ate") or ""
    q = AjusteInventario.query.filter_by(tipo="baixa", status="aplicado")
    if de:
        try:
            q = q.filter(AjusteInventario.criado_em >= datetime.strptime(de, "%Y-%m-%d"))
        except ValueError:
            pass
    if ate:
        try:
            fim = datetime.strptime(ate, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(AjusteInventario.criado_em <= fim)
        except ValueError:
            pass
    perdas = q.order_by(AjusteInventario.criado_em.desc()).all()
    total = sum(abs(a.diferenca or 0) for a in perdas)
    return render_template("almox/relatorio_perdas.html", perdas=perdas, total=total, de=de, ate=ate)


# ==================== COLETOR — BLOCO 5: MOVER MATERIAL ====================
@almox_bp.route("/coletor/api/estoque/<int:produto_id>/<int:loc_id>")
@modulo_required
def coletor_api_estoque(produto_id, loc_id):
    if not _aba_material_qualquer(): return _nega_aba("material")
    """Quantidade de um item numa prateleira específica (para saber se há 2+ ao mover)."""
    p = db.session.get(ProdutoAlmox, produto_id)
    if not p:
        return jsonify(ok=False, erro="Item não encontrado."), 404
    linha = p.estoque_em(loc_id)
    return jsonify(ok=True, quantidade=(linha.quantidade or 0) if linha else 0, nome=p.nome)


@almox_bp.route("/coletor/api/mover", methods=["POST"])
@modulo_required
def coletor_api_mover_v2():
    if not _aba_coletor_ok("mover"): return _nega_aba("mover")
    """MOVER material entre localizadores. Fase 2 (destinar).
    payload: {senha, destino_loc_id, destino_cod, ciente, itens:[{produto_id, origem_loc_id, origem_cod, qtd}]}
    Se algum item já existir em OUTRO localizador (além do destino), retorna needs_confirm até 'ciente'."""
    data = request.get_json(silent=True) or {}
    if not _current_pode_conferir(data.get("senha")):
        return jsonify(ok=False, erro="Senha do almoxarife inválida."), 403
    destino_id = data.get("destino_loc_id")
    destino_cod = data.get("destino_cod") or ""
    itens = data.get("itens") or []
    if not destino_id or not itens:
        return jsonify(ok=False, erro="Informe destino e itens."), 400
    # valida quantidade na origem
    faltas = []
    for it in itens:
        p = db.session.get(ProdutoAlmox, it.get("produto_id") or 0)
        if not p:
            continue
        linha = p.estoque_em(it.get("origem_loc_id"))
        disp = (linha.quantidade or 0) if linha else 0
        if _num(it.get("qtd"), 0) > disp:
            faltas.append(f"{p.nome} em {it.get('origem_cod')} (tem {disp:g})")
    if faltas:
        return jsonify(ok=False, erro="Quantidade indisponível na origem: " + "; ".join(faltas)), 400
    # aviso: item já existe em outro localizador (além do destino)
    if not data.get("ciente"):
        conflitos = []
        for it in itens:
            p = db.session.get(ProdutoAlmox, it.get("produto_id") or 0)
            if not p:
                continue
            outros = [l.local_cod for l in p.linhas_estoque()
                      if l.localizador_id and l.localizador_id != int(destino_id)
                      and l.localizador_id != it.get("origem_loc_id") and (l.quantidade or 0) > 0]
            if outros:
                conflitos.append(f"{p.nome}: já existe em {', '.join(sorted(set(outros)))}")
        if conflitos:
            return jsonify(ok=False, needs_confirm=True,
                           aviso="Este(s) item(ns) já existe(m) em outro localizador. Confirme para mover assim mesmo.",
                           detalhes=conflitos)
    # aplica
    feitas = []
    for it in itens:
        p = db.session.get(ProdutoAlmox, it.get("produto_id") or 0)
        qn = _num(it.get("qtd"), 0)
        if not p or qn <= 0:
            continue
        p.ajustar_estoque(it.get("origem_loc_id"), -qn)
        p.ajustar_estoque(int(destino_id), qn)
        db.session.add(MovimentacaoMaterial(produto_id=p.id, produto_nome=p.nome, tipo="movimentacao",
                       quantidade=qn, saldo_apos=p.saldo, local_de=it.get("origem_cod"),
                       local_para=destino_cod, operador_id=_op_id(), obs="Coletor/mover"))
        feitas.append(f"{qn:g} {p.unidade} de {p.nome}: {it.get('origem_cod')} → {destino_cod}")
    _log("Coletor", f"Mover para {destino_cod}: " + "; ".join(feitas))
    db.session.commit()
    return jsonify(ok=True, resumo=feitas)


# ==================== COLETOR — BLOCO 6: AJUSTE DE INSTÂNCIAS ====================
@almox_bp.route("/coletor/api/instancias/<int:produto_id>/<int:loc_id>")
@modulo_required
def coletor_api_instancias(produto_id, loc_id):
    if not _aba_coletor_ok("inventario"): return _nega_aba("inventario")
    """Instâncias (unidades/TAG) de um item num localizador + os campos opcionais do cadastro-raiz."""
    p = db.session.get(ProdutoAlmox, produto_id)
    if not p:
        return jsonify(ok=False, erro="Item não encontrado."), 404
    insts = InstanciaItem.query.filter_by(produto_id=produto_id, localizador_id=loc_id).all()
    linha = p.estoque_em(loc_id)
    saldo_loc = (linha.quantidade or 0) if linha else 0
    def _d(x):
        return x.isoformat() if x else ""
    lst = [{"id": i.id, "tag": i.tag or "(sem TAG)", "ca": i.ca or "", "lote": i.lote or "",
            "validade": _d(i.validade), "validade_calib": _d(i.validade_calib),
            "quantidade": i.quantidade or 1} for i in insts]
    return jsonify(ok=True, instancias=lst, opcionais=p.opcionais_ativos, saldo_loc=saldo_loc, nome=p.nome)


@almox_bp.route("/coletor/api/instancia-salvar", methods=["POST"])
@modulo_required
def coletor_api_instancia_salvar():
    if not _aba_coletor_ok("inventario"): return _nega_aba("inventario")
    """Ajusta os dados de N unidades de um item (não mexe na quantidade de estoque).
    payload: {produto_id, loc_id, instancia_id (ou null), quantidade, campos:{tag,ca,validade,validade_calib,lote}, senha}"""
    from datetime import datetime as _dt
    data = request.get_json(silent=True) or {}
    if not _current_pode_conferir(data.get("senha")):
        return jsonify(ok=False, erro="Senha do almoxarife inválida."), 403
    p = db.session.get(ProdutoAlmox, data.get("produto_id") or 0)
    if not p:
        return jsonify(ok=False, erro="Item inválido."), 400
    loc_id = data.get("loc_id")
    n = int(_num(data.get("quantidade"), 1))
    if n < 1:
        return jsonify(ok=False, erro="Quantidade inválida."), 400
    campos = data.get("campos") or {}
    def _pdate(s):
        try:
            return _dt.strptime(s, "%Y-%m-%d").date() if s else None
        except ValueError:
            return None
    def _aplica(inst):
        if "tag" in campos: inst.tag = (campos.get("tag") or "").strip() or None
        if "ca" in campos: inst.ca = (campos.get("ca") or "").strip() or None
        if "lote" in campos: inst.lote = (campos.get("lote") or "").strip() or None
        if "validade" in campos: inst.validade = _pdate(campos.get("validade"))
        if "validade_calib" in campos: inst.validade_calib = _pdate(campos.get("validade_calib"))

    inst_id = data.get("instancia_id")
    if inst_id:
        inst = db.session.get(InstanciaItem, int(inst_id))
        if not inst:
            return jsonify(ok=False, erro="Instância não encontrada."), 404
        if n >= (inst.quantidade or 1):
            _aplica(inst)                       # ajusta todas as unidades daquela instância
        else:
            inst.quantidade = (inst.quantidade or 1) - n   # separa N unidades numa nova instância
            nova = InstanciaItem(produto_id=p.id, localizador_id=loc_id, quantidade=n)
            _aplica(nova); db.session.add(nova)
    else:
        # cria uma instância nova a partir do estoque do localizador
        linha = p.estoque_em(loc_id)
        disp = (linha.quantidade or 0) if linha else 0
        if n > disp and disp > 0:
            n = int(disp)
        nova = InstanciaItem(produto_id=p.id, localizador_id=loc_id, quantidade=n)
        _aplica(nova); db.session.add(nova)
    _log("Coletor", f"Ajuste de instância: {n} unidade(s) de {p.nome}")
    db.session.commit()
    return jsonify(ok=True, resumo=[f"Ajuste aplicado em {n} unidade(s) de {p.nome}"])


# ==================== ADMINISTRATIVO (desktop): NOTAS FISCAIS + NOTIFICAÇÕES ====================
import json as _json
import xml.etree.ElementTree as _ET


@almox_bp.route("/notificacoes")
@_guard("is_admin")
def notificacoes():
    itens = NotificacaoAlmox.query.order_by(NotificacaoAlmox.lida, NotificacaoAlmox.criado_em.desc()).all()
    return render_template("almox/notificacoes.html", itens=itens)


@almox_bp.route("/notificacoes/<int:nid>/lida", methods=["POST"])
@_guard("is_admin")
def notificacao_lida(nid):
    n = NotificacaoAlmox.query.get_or_404(nid)
    n.lida = True
    db.session.commit()
    return redirect(request.referrer or url_for("almox.notificacoes"))


@almox_bp.route("/administrativo")
@_guard("is_admin")
def administrativo():
    n_notif = NotificacaoAlmox.query.filter_by(lida=False).count()
    n_classificar = NotaFiscalAlmox.query.filter(NotaFiscalAlmox.classificacao.is_(None)).count()
    return render_template("almox/administrativo.html", n_notif=n_notif, n_classificar=n_classificar)


def _parse_nfe_xml(conteudo_bytes):
    """Extrai dados de uma NF-e (XML padrão SEFAZ). Retorna dict ou None."""
    try:
        txt = conteudo_bytes.decode("utf-8", errors="ignore")
        # remove namespace para simplificar as buscas
        txt = txt.replace('xmlns="http://www.portalfiscal.inf.br/nfe"', "")
        root = _ET.fromstring(txt)
    except Exception:
        return None
    def _find(path):
        el = root.find(path)
        return el.text if el is not None else None
    numero = _find(".//ide/nNF")
    dh = _find(".//ide/dhEmi") or _find(".//ide/dEmi")
    data_em = None
    if dh:
        try:
            data_em = datetime.strptime(dh[:10], "%Y-%m-%d").date()
        except ValueError:
            data_em = None
    fornecedor = _find(".//emit/xNome")
    valor = _find(".//total/ICMSTot/vNF")
    itens = []
    for det in root.findall(".//det"):
        prod = det.find("prod")
        if prod is not None:
            itens.append({"nome": (prod.findtext("xProd") or "").strip(),
                          "qtd": prod.findtext("qCom"), "valor": prod.findtext("vUnCom")})
    if not (numero or fornecedor):
        return None
    return {"numero": numero, "fornecedor": fornecedor, "valor": _num(valor, 0),
            "data_emissao": data_em, "itens": itens}


@almox_bp.route("/administrativo/notas")
@_guard("is_admin")
def adm_notas():
    a_classificar = (NotaFiscalAlmox.query.filter(NotaFiscalAlmox.classificacao.is_(None))
                     .order_by(NotaFiscalAlmox.criado_em.desc()).all())
    classificadas = (NotaFiscalAlmox.query.filter(NotaFiscalAlmox.classificacao.isnot(None))
                     .order_by(NotaFiscalAlmox.criado_em.desc()).limit(50).all())
    return render_template("almox/adm_notas.html", a_classificar=a_classificar, classificadas=classificadas)


@almox_bp.route("/administrativo/notas/importar", methods=["POST"])
@_guard("is_admin")
def adm_notas_importar():
    arq = request.files.get("arquivo")
    if not arq or not arq.filename:
        flash("Selecione um arquivo XML ou PDF da nota.", "danger")
        return redirect(url_for("almox.adm_notas"))
    nome = arq.filename.lower()
    dados = arq.read()
    info = None
    if nome.endswith(".xml"):
        info = _parse_nfe_xml(dados)
        if not info:
            flash("Não consegui ler este XML como NF-e. Confira o arquivo.", "warning")
            return redirect(url_for("almox.adm_notas"))
    else:
        # PDF: leitura estruturada é limitada; cria a nota para preenchimento/conferência manual.
        flash("PDF recebido. A leitura automática de PDF é limitada — confira e complete os campos.", "info")
        info = {"numero": "", "fornecedor": "", "valor": 0, "data_emissao": None, "itens": []}
    nf = NotaFiscalAlmox(numero=(info.get("numero") or "").strip(),
                         fornecedor_nome=(info.get("fornecedor") or "").strip().upper() or None,
                         valor=info.get("valor") or 0, data_emissao=info.get("data_emissao"),
                         itens_json=_json.dumps(info.get("itens") or [], ensure_ascii=False),
                         origem="importada")
    db.session.add(nf); db.session.flush()
    db.session.add(NotificacaoAlmox(tipo="classificar_nf", titulo="Classificar OPEX/CAPEX",
                   texto=f"Nota {nf.rotulo} importada. Classifique OPEX ou CAPEX.", ref_id=nf.id))
    _log("Administrativo", f"NF importada: {nf.rotulo}")
    db.session.commit()
    return redirect(url_for("almox.adm_nota_ver", nid=nf.id))


@almox_bp.route("/administrativo/notas/<int:nid>")
@_guard("is_admin")
def adm_nota_ver(nid):
    nf = NotaFiscalAlmox.query.get_or_404(nid)
    itens = []
    try:
        itens = _json.loads(nf.itens_json) if nf.itens_json else []
    except Exception:
        itens = []
    return render_template("almox/adm_nota_ver.html", nf=nf, itens=itens)


@almox_bp.route("/administrativo/notas/<int:nid>/classificar", methods=["POST"])
@_guard("is_admin")
def adm_nota_classificar(nid):
    nf = NotaFiscalAlmox.query.get_or_404(nid)
    # permite completar campos (útil no caso do PDF)
    nf.numero = (request.form.get("numero") or nf.numero or "").strip()
    forn = (request.form.get("fornecedor") or "").strip().upper()
    if forn:
        nf.fornecedor_nome = forn
    if request.form.get("valor"):
        nf.valor = _num(request.form.get("valor"), nf.valor or 0)
    nf.ordem_compra = (request.form.get("ordem_compra") or "").strip() or None
    classe = (request.form.get("classificacao") or "").lower()
    if classe not in ("opex", "capex"):
        flash("Escolha OPEX ou CAPEX.", "danger")
        return redirect(url_for("almox.adm_nota_ver", nid=nid))
    nf.classificacao = classe
    # marca as notificações dessa nota como lidas
    for n in NotificacaoAlmox.query.filter_by(tipo="classificar_nf", ref_id=nf.id, lida=False).all():
        n.lida = True
    _log("Administrativo", f"NF {nf.rotulo} classificada como {classe.upper()}"
         + (f" · OC {nf.ordem_compra}" if nf.ordem_compra else ""))
    db.session.commit()
    flash(f"Nota classificada como {classe.upper()}.", "success")
    return redirect(url_for("almox.adm_notas"))


# ==================== "VER COMO PERFIL" (admin master) ====================
from flask import session as _session

_PERM_MAP = {
    "is_admin": "perm_total",
    "pode_almox_modulo": "perm_modulo_almox",
    "is_almox": "perm_modulo_almox",
    "pode_chaves": "perm_chaves",
    "pode_extintores": "perm_extintores",
    "pode_colaboradores": "perm_colaboradores",
    "pode_solicitar": "perm_solicitar",
}


def _ver_como_nome():
    """Nome do perfil que o master está simulando (ou None). Só master; só perfil ativo."""
    try:
        if not getattr(current_user, "is_master", False):
            return None
        nome = _session.get("ver_como_perfil")
        if not nome:
            return None
        perfil = PapelColaborador.query.filter(
            db.func.upper(PapelColaborador.nome) == nome.upper(),
            PapelColaborador.ativo.is_(True)).first()
        return perfil.nome if perfil else None
    except Exception:
        return None


def _perms_simuladas():
    nome = _ver_como_nome()
    if not nome:
        return None
    perfil = PapelColaborador.query.filter(db.func.upper(PapelColaborador.nome) == nome.upper()).first()
    return set(perfil.lista_tarefas) if perfil else set()


def _efetivo(prop):
    """Valor efetivo de uma permissão/tarefa. Se simulando um perfil, usa as tarefas dele
    (perm_from_tasks); senão usa a propriedade do usuário, e se não existir, cai em tem_tarefa
    (para nomes de tarefa como ext_cadastrar, ext_desativar, chave_desativar). is_master nunca simula."""
    if prop == "is_master":
        return bool(getattr(current_user, "is_master", False))
    perms = _perms_simuladas()
    if perms is not None:
        from .models import perm_from_tasks
        return perm_from_tasks(perms, prop)
    if hasattr(current_user, prop):
        return bool(getattr(current_user, prop))
    tem = getattr(current_user, "tem_tarefa", None)
    return bool(tem(prop)) if tem else False


def _bloqueado_ver_como():
    return render_template("almox/ver_como_bloqueado.html", perfil=_ver_como_nome())


@almox_bp.route("/ver-como/ativar", methods=["POST"])
@login_required
def ver_como_ativar():
    if not getattr(current_user, "is_master", False):
        abort(403)
    nome = (request.form.get("perfil") or "").strip()
    perfil = PapelColaborador.query.filter(
        db.func.upper(PapelColaborador.nome) == nome.upper(), PapelColaborador.ativo.is_(True)).first()
    if not perfil:
        flash("Perfil inválido ou inativo.", "danger")
        return redirect(request.referrer or url_for("almox.papeis"))
    _session["ver_como_perfil"] = perfil.nome
    flash(f"Você está vendo como o perfil {perfil.nome}.", "info")
    return redirect(url_for("almox.home"))


@almox_bp.route("/ver-como/sair")
@login_required
def ver_como_sair():
    """Sai da simulação. Não depende de nenhuma permissão do perfil simulado (anti-travamento)."""
    _session.pop("ver_como_perfil", None)
    flash("Você voltou à sua visão normal.", "success")
    return redirect(url_for("almox.home"))


@almox_bp.app_context_processor
def _inject_ver_como():
    """Expõe 'perm' (permissões efetivas) e 'ver_como_nome' para os templates/menu."""
    try:
        if not current_user.is_authenticated:
            return {}
    except Exception:
        return {}
    nome = _ver_como_nome()

    class _Perm:
        pass
    p = _Perm()
    for prop in ("is_admin", "is_master", "is_almox", "pode_almox_modulo", "pode_chaves",
                 "pode_extintores", "pode_material", "pode_locais", "pode_relatorios",
                 "pode_coletor", "pode_criar_solicitacao", "pode_ver_solicitacoes",
                 "pode_colaboradores", "pode_solicitar",
                 "ext_cadastrar", "ext_desativar", "chave_desativar"):
        setattr(p, prop, _efetivo(prop))
    return {"perm": p, "ver_como_nome": nome}
